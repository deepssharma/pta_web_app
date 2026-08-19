"""
Tests for budget_io.py — the Excel-template budget config.
"""
import openpyxl
import pytest

from pta_treasurer.budget_io import (
    SHEET_NAMES, actuals_total_for_item, add_item, apply_dynamic_last_year,
    apply_edit, describe_edit, generate_template, load_budget,
    map_actuals_to_budget_items, map_qb_category, merge_actuals_into_budget,
    move_item, remove_item, rename_item, set_budget_amount, unmap_qb_category,
)


# ── generate_template ─────────────────────────────────────────────────────

def test_generate_template_creates_both_sheets(tmp_path):
    path = tmp_path / 'budget.xlsx'
    generate_template(path)
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == list(SHEET_NAMES)


def test_generate_template_has_column_headers(tmp_path):
    path = tmp_path / 'budget.xlsx'
    generate_template(path)
    wb = openpyxl.load_workbook(path)
    header_row = [c.value for c in wb['Income Budget'][2]]
    assert header_row == [
        'Section', 'Item', 'QuickBooks Category Name(s)',
        'Last Year Actual', 'This Year Budget',
    ]


# ── load_budget ───────────────────────────────────────────────────────────

def _fill_budget(path, income_rows, expense_rows):
    generate_template(path)
    wb = openpyxl.load_workbook(path)
    for sheet_name, rows in [('Income Budget', income_rows), ('Expense Budget', expense_rows)]:
        ws = wb[sheet_name]
        for i, row in enumerate(rows, start=3):
            for ci, val in enumerate(row, start=1):
                ws.cell(row=i, column=ci, value=val)
    wb.save(path)


def test_load_budget_parses_rows(tmp_path):
    path = tmp_path / 'budget.xlsx'
    _fill_budget(
        path,
        income_rows=[('Fundraising', 'Book Fair', '', 9118.36, 500.0)],
        expense_rows=[('Admin/General', 'Bank Services', '', 234.94, 200.0)],
    )
    income_budget, expense_budget, qb_map = load_budget(path)
    assert income_budget == {'Fundraising': {'Book Fair': (9118.36, 500.0)}}
    assert expense_budget == {'Admin/General': {'Bank Services': (234.94, 200.0)}}


def test_load_budget_builds_qb_map_with_multiple_names(tmp_path):
    path = tmp_path / 'budget.xlsx'
    _fill_budget(
        path,
        income_rows=[],
        expense_rows=[
            ('Admin/General', 'Accounting', 'Accounting Expense (Quickbooks), QB Fees', 0, 1300.0),
        ],
    )
    _, _, qb_map = load_budget(path)
    assert qb_map == {
        'Accounting Expense (Quickbooks)': 'Accounting',
        'QB Fees': 'Accounting',
    }


def test_load_budget_skips_blank_rows(tmp_path):
    path = tmp_path / 'budget.xlsx'
    _fill_budget(
        path,
        income_rows=[('Fundraising', 'Book Fair', '', 100.0, 200.0), (None, None, None, None, None)],
        expense_rows=[],
    )
    income_budget, _, _ = load_budget(path)
    assert list(income_budget['Fundraising'].keys()) == ['Book Fair']


# ── merge_actuals_into_budget ─────────────────────────────────────────────

def test_merge_direct_name_match():
    budget = {'Program Expense': {'Picnic': (0.0, 200.0)}}
    actuals = {'Picnic': [181.58] + [0.0] * 11}
    merged = merge_actuals_into_budget(budget, actuals, qb_to_budget_map={})
    assert merged == {'Program Expense': {'Picnic': (0.0, 200.0, [181.58] + [0.0] * 11)}}


def test_merge_maps_qb_category_to_budget_item():
    budget = {'Admin/General': {'Accounting': (0.0, 1300.0)}}
    actuals = {'Accounting Expense (Quickbooks)': [1257.76] + [0.0] * 11}
    qb_map = {'Accounting Expense (Quickbooks)': 'Accounting'}
    merged = merge_actuals_into_budget(budget, actuals, qb_map)
    assert merged['Admin/General']['Accounting'][2][0] == 1257.76


def test_merge_sums_multiple_qb_categories_into_one_item():
    budget = {'Admin/General': {'Accounting': (0.0, 1300.0)}}
    actuals = {
        'Accounting Expense (Quickbooks)': [1000.0] + [0.0] * 11,
        'QB Fees': [50.0] + [0.0] * 11,
    }
    qb_map = {'Accounting Expense (Quickbooks)': 'Accounting', 'QB Fees': 'Accounting'}
    merged = merge_actuals_into_budget(budget, actuals, qb_map)
    assert merged['Admin/General']['Accounting'][2][0] == 1050.0


def test_merge_unmatched_qb_category_goes_to_other_section():
    budget = {'Admin/General': {'Accounting': (0.0, 1300.0)}}
    actuals = {'Mystery Category': [42.0] + [0.0] * 11}
    merged = merge_actuals_into_budget(budget, actuals, qb_to_budget_map={})
    assert merged['Other (from QuickBooks)']['Mystery Category'] == (0.0, 0.0, [42.0] + [0.0] * 11)


def test_merge_item_with_no_actuals_gets_zeros():
    budget = {'Fundraising': {'Book Fair': (9118.36, 500.0)}}
    merged = merge_actuals_into_budget(budget, actuals_dict={}, qb_to_budget_map={})
    assert merged['Fundraising']['Book Fair'] == (9118.36, 500.0, [0.0] * 12)


# ── map_actuals_to_budget_items ──────────────────────────────────────────

def test_map_actuals_translates_qb_name_to_budget_item():
    actuals = {'Accounting Expense (Quickbooks)': [1257.76] + [0.0] * 11}
    qb_map = {'Accounting Expense (Quickbooks)': 'Accounting'}
    mapped = map_actuals_to_budget_items(actuals, qb_map)
    assert mapped == {'Accounting': [1257.76] + [0.0] * 11}


def test_map_actuals_sums_multiple_qb_categories_into_one_item():
    actuals = {
        'Accounting Expense (Quickbooks)': [1000.0] + [0.0] * 11,
        'QB Fees': [50.0] + [0.0] * 11,
    }
    qb_map = {'Accounting Expense (Quickbooks)': 'Accounting', 'QB Fees': 'Accounting'}
    mapped = map_actuals_to_budget_items(actuals, qb_map)
    assert mapped['Accounting'][0] == 1050.0


def test_map_actuals_unmapped_category_keeps_its_own_name():
    actuals = {'Mystery Category': [42.0] + [0.0] * 11}
    mapped = map_actuals_to_budget_items(actuals, qb_to_budget_map={})
    assert mapped == {'Mystery Category': [42.0] + [0.0] * 11}


# ── apply_dynamic_last_year ──────────────────────────────────────────────

def test_apply_dynamic_last_year_no_prior_history_keeps_static_value():
    budget = {'Program Expense': {'Picnic': (200.0, 250.0)}}
    result = apply_dynamic_last_year(budget, prior_actuals_mapped={})
    assert result == budget


def test_apply_dynamic_last_year_replaces_with_prior_actual():
    budget = {'Program Expense': {'Picnic': (200.0, 250.0)}}
    prior = {'Picnic': [181.58] + [0.0] * 11}
    result = apply_dynamic_last_year(budget, prior)
    assert result == {'Program Expense': {'Picnic': (181.58, 250.0)}}


def test_apply_dynamic_last_year_sums_all_12_months():
    budget = {'Program Expense': {'Picnic': (200.0, 250.0)}}
    prior = {'Picnic': [100.0, 50.0] + [0.0] * 10}
    result = apply_dynamic_last_year(budget, prior)
    assert result['Program Expense']['Picnic'][0] == 150.0


def test_apply_dynamic_last_year_item_with_no_prior_actual_becomes_zero():
    # Some prior-year history exists (for a DIFFERENT item), but this
    # specific item has no matching prior-year actual at all -- that's a
    # genuinely zero prior year for it, not a reason to keep the stale
    # static value.
    budget = {'Program Expense': {'Picnic': (200.0, 250.0)}}
    prior = {'Some Other Item': [500.0] + [0.0] * 11}
    result = apply_dynamic_last_year(budget, prior)
    assert result == {'Program Expense': {'Picnic': (0.0, 250.0)}}


# ── actuals_total_for_item ────────────────────────────────────────────────

def test_actuals_total_for_item_sums_across_12_months():
    merged = {'Fundraising': {'Book Fair': (100.0, 500.0, [10.0, 20.0] + [0.0] * 10)}}
    assert actuals_total_for_item(merged, 'Book Fair') == 30.0


def test_actuals_total_for_item_missing_item_is_zero():
    merged = {'Fundraising': {'Book Fair': (100.0, 500.0, [10.0] + [0.0] * 11)}}
    assert actuals_total_for_item(merged, 'Nonexistent') == 0.0


# ── add_item / remove_item / move_item / rename_item / map_qb_category /
#    set_budget_amount ────────────────────────────────────────────────────

def _budget_path(tmp_path, income_rows=(), expense_rows=()):
    path = tmp_path / 'budget.xlsx'
    _fill_budget(path, list(income_rows), list(expense_rows))
    return path


def test_add_item_appends_new_row(tmp_path):
    path = _budget_path(tmp_path, income_rows=[('Fundraising', 'Book Fair', '', 100.0, 200.0)])
    add_item(path, 'Income Budget', 'Program', 'Talent Show',
              qb_names=['Talent Show Income'], budget=750.0)
    income_budget, _, qb_map = load_budget(path)
    assert income_budget['Program']['Talent Show'] == (0.0, 750.0)
    assert qb_map['Talent Show Income'] == 'Talent Show'
    # existing row untouched
    assert income_budget['Fundraising']['Book Fair'] == (100.0, 200.0)


def test_add_item_rejects_duplicate(tmp_path):
    path = _budget_path(tmp_path, income_rows=[('Fundraising', 'Book Fair', '', 100.0, 200.0)])
    with pytest.raises(ValueError, match='already exists'):
        add_item(path, 'Income Budget', 'Fundraising', 'Book Fair')


def test_add_item_invalid_sheet_raises(tmp_path):
    path = _budget_path(tmp_path)
    with pytest.raises(ValueError, match='sheet must be one of'):
        add_item(path, 'Not A Real Sheet', 'Section', 'Item')


def test_add_item_writes_backup_file(tmp_path):
    path = _budget_path(tmp_path, income_rows=[('Fundraising', 'Book Fair', '', 100.0, 200.0)])
    add_item(path, 'Income Budget', 'Program', 'Talent Show')
    backup = path.with_name(f'{path.stem}.bak{path.suffix}')
    assert backup.exists()
    old_income, _, _ = load_budget(backup)
    assert 'Talent Show' not in old_income.get('Program', {})


def test_remove_item_deletes_row(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[
        ('Admin', 'Bank Services', '', 50.0, 100.0),
        ('Admin', 'Accounting', '', 650.0, 650.0),
    ])
    remove_item(path, 'Expense Budget', 'Bank Services')
    _, expense_budget, _ = load_budget(path)
    assert 'Bank Services' not in expense_budget.get('Admin', {})
    assert expense_budget['Admin']['Accounting'] == (650.0, 650.0)


def test_remove_item_not_found_raises(tmp_path):
    path = _budget_path(tmp_path)
    with pytest.raises(ValueError, match='not found'):
        remove_item(path, 'Expense Budget', 'Nonexistent')


def test_remove_item_refuses_when_actuals_present(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[('Admin', 'Bank Services', '', 50.0, 100.0)])
    with pytest.raises(ValueError, match=r'\$234\.94'):
        remove_item(path, 'Expense Budget', 'Bank Services', current_actuals_total=234.94)
    # untouched
    _, expense_budget, _ = load_budget(path)
    assert 'Bank Services' in expense_budget['Admin']


def test_remove_item_force_overrides_actuals_guard(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[('Admin', 'Bank Services', '', 50.0, 100.0)])
    remove_item(path, 'Expense Budget', 'Bank Services',
                current_actuals_total=234.94, force=True)
    _, expense_budget, _ = load_budget(path)
    assert 'Bank Services' not in expense_budget.get('Admin', {})


def test_remove_item_skips_guard_when_actuals_total_is_none(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[('Admin', 'Bank Services', '', 50.0, 100.0)])
    remove_item(path, 'Expense Budget', 'Bank Services', current_actuals_total=None)
    _, expense_budget, _ = load_budget(path)
    assert 'Bank Services' not in expense_budget.get('Admin', {})


def test_move_item_rekeys_section(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[('Programs', 'Field Day', '', 0.0, 200.0)])
    move_item(path, 'Expense Budget', 'Field Day', 'Grad Class Activities')
    _, expense_budget, _ = load_budget(path)
    assert 'Field Day' not in expense_budget.get('Programs', {})
    assert expense_budget['Grad Class Activities']['Field Day'] == (0.0, 200.0)


def test_move_item_not_found_raises(tmp_path):
    path = _budget_path(tmp_path)
    with pytest.raises(ValueError, match='not found'):
        move_item(path, 'Expense Budget', 'Nonexistent', 'Somewhere')


def test_rename_item_updates_name_and_keeps_qb_mapping(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[
        ('Admin', 'Accounting', 'Accounting Expense (Quickbooks)', 650.0, 650.0),
    ])
    rename_item(path, 'Expense Budget', 'Accounting', 'Accounting Quickbooks')
    _, expense_budget, qb_map = load_budget(path)
    assert 'Accounting' not in expense_budget.get('Admin', {})
    assert expense_budget['Admin']['Accounting Quickbooks'] == (650.0, 650.0)
    assert qb_map['Accounting Expense (Quickbooks)'] == 'Accounting Quickbooks'


def test_rename_item_rejects_collision_with_existing_item(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[
        ('Admin', 'Accounting', '', 650.0, 650.0),
        ('Admin', 'Bank Services', '', 50.0, 100.0),
    ])
    with pytest.raises(ValueError, match='already exists'):
        rename_item(path, 'Expense Budget', 'Accounting', 'Bank Services')


def test_map_qb_category_appends_to_existing_list(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[
        ('Admin', 'Accounting', 'Accounting Expense (Quickbooks)', 650.0, 650.0),
    ])
    map_qb_category(path, 'Expense Budget', 'Accounting', 'QB Fees')
    _, _, qb_map = load_budget(path)
    assert qb_map['Accounting Expense (Quickbooks)'] == 'Accounting'
    assert qb_map['QB Fees'] == 'Accounting'


def test_map_qb_category_is_idempotent(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[
        ('Admin', 'Accounting', 'Accounting Expense (Quickbooks)', 650.0, 650.0),
    ])
    map_qb_category(path, 'Expense Budget', 'Accounting', 'Accounting Expense (Quickbooks)')
    wb = openpyxl.load_workbook(path)
    row = next(r for r in wb['Expense Budget'].iter_rows(min_row=3, values_only=True) if r[1] == 'Accounting')
    assert row[2] == 'Accounting Expense (Quickbooks)'


def test_unmap_qb_category_removes_from_list(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[
        ('Admin', 'Accounting', 'Accounting Expense (Quickbooks), QB Fees', 650.0, 650.0),
    ])
    unmap_qb_category(path, 'Expense Budget', 'Accounting', 'QB Fees')
    _, _, qb_map = load_budget(path)
    assert qb_map == {'Accounting Expense (Quickbooks)': 'Accounting'}


def test_unmap_qb_category_moving_between_items(tmp_path):
    # The real-world case this exists for: a QB category was mapped to
    # the wrong budget item and needs to move to the right one.
    path = _budget_path(tmp_path, expense_rows=[
        ('Donations', 'Staff Shirts', '5th grade T-shirts', 191.78, 125.0),
        ('Grad Class Activities', 'Grade T-Shirts', '', 1494.25, 800.0),
    ])
    unmap_qb_category(path, 'Expense Budget', 'Staff Shirts', '5th grade T-shirts')
    map_qb_category(path, 'Expense Budget', 'Grade T-Shirts', '5th grade T-shirts')
    _, _, qb_map = load_budget(path)
    assert qb_map == {'5th grade T-shirts': 'Grade T-Shirts'}


def test_unmap_qb_category_noop_when_not_present(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[
        ('Admin', 'Accounting', 'Accounting Expense (Quickbooks)', 650.0, 650.0),
    ])
    unmap_qb_category(path, 'Expense Budget', 'Accounting', 'Nonexistent Category')
    _, _, qb_map = load_budget(path)
    assert qb_map == {'Accounting Expense (Quickbooks)': 'Accounting'}


def test_set_budget_amount_updates_this_year_column(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[('Admin', 'Accounting', '', 650.0, 650.0)])
    set_budget_amount(path, 'Expense Budget', 'Accounting', 800.0)
    _, expense_budget, _ = load_budget(path)
    assert expense_budget['Admin']['Accounting'] == (650.0, 800.0)


def test_set_budget_amount_not_found_raises(tmp_path):
    path = _budget_path(tmp_path)
    with pytest.raises(ValueError, match='not found'):
        set_budget_amount(path, 'Expense Budget', 'Nonexistent', 100.0)


# ── describe_edit ──────────────────────────────────────────────────────────

def test_describe_edit_add_item():
    text = describe_edit({'action': 'add_item', 'sheet': 'Expense Budget',
                           'section': 'Programs', 'item': 'Robotics Club', 'budget': 500.0})
    assert 'Robotics Club' in text and 'Programs' in text and '$500.00' in text


def test_describe_edit_move_item():
    text = describe_edit({'action': 'move_item', 'sheet': 'Expense Budget',
                           'item': 'Field Day', 'to_section': 'Grad Class Activities'})
    assert 'Field Day' in text and 'Grad Class Activities' in text


def test_describe_edit_unknown_action_raises():
    with pytest.raises(ValueError, match='Unknown edit action'):
        describe_edit({'action': 'nonsense'})


# ── apply_edit ─────────────────────────────────────────────────────────────

def test_apply_edit_dispatches_add_item(tmp_path):
    path = _budget_path(tmp_path)
    apply_edit(path, {'action': 'add_item', 'sheet': 'Income Budget',
                       'section': 'Program', 'item': 'Talent Show', 'budget': 750.0})
    income_budget, _, _ = load_budget(path)
    assert income_budget['Program']['Talent Show'] == (0.0, 750.0)


def test_apply_edit_dispatches_remove_item_with_actuals_guard(tmp_path):
    path = _budget_path(tmp_path, expense_rows=[('Admin', 'Bank Services', '', 50.0, 100.0)])
    with pytest.raises(ValueError, match=r'\$10\.00'):
        apply_edit(path, {'action': 'remove_item', 'sheet': 'Expense Budget', 'item': 'Bank Services'},
                   current_actuals_total=10.0)


def test_apply_edit_unknown_action_raises(tmp_path):
    path = _budget_path(tmp_path)
    with pytest.raises(ValueError, match='Unknown edit action'):
        apply_edit(path, {'action': 'nonsense'})
