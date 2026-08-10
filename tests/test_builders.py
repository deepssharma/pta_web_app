"""
Tests for builders.py Excel sheet builder functions.
Uses mock data — no real files needed.
"""
import pytest
import openpyxl
from pta_treasurer.builders import (build_treasurer, build_budget, build_givebacks,
                      build_manifest, build_ytd_summary, FISCAL_MONTHS, GOLD_FILL, RED_FILL,
                      build_credits_sheet, build_debits_sheet, build_memberhub_summary_sheet)


# ── Mock data ─────────────────────────────────────────────────────────────────

MOCK_QRB = {
    'period':        'July 1-31, 2025',
    'income':        {},
    'income_total':  0.0,
    'expenses':      {'Accounting Quickbooks': 1257.76, '6th Grade Events': 181.58},
    'expense_total': 1439.34,
    'net_income':    -1439.34,
    'transactions':  [],
}

MOCK_BANK = {
    'period':            'July 01, 2025 through July 31, 2025',
    'account':           '4346',
    'beginning_balance': 32630.10,
    'ending_balance':    31190.76,
    'total_deposits':    0.0,
    'total_checks':      181.58,
    'total_withdrawals': 1257.76,
    'total_fees':        0.0,
    'deposits':          [],
    'checks':            [{'check_no': '1077', 'date': '07/01', 'amount': 181.58}],
    'withdrawals':       [],
    'fees':              [],
    'daily_balances':    {'07/31': 31190.76},
    'source_file':       'Chase_july_2025.pdf',
}

MOCK_GIVEBACKS = [
    {'item': 'Shop to Give Donation', 'category': '',
     'count': 1, 'total': 95.14, 'source_file': 'givebacks_july.csv'},
    {'item': 'Teacher/Staff', 'category': 'Memberships',
     'count': 1, 'total': 15.0,  'source_file': 'givebacks_july.csv'},
]


# ── Treasurer Report tests ────────────────────────────────────────────────────

def test_build_treasurer_creates_sheet():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA')
    assert ws['A1'].value == 'SETAUKET SCHOOL PTA'

def test_build_treasurer_has_org_name():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Test PTA')
    assert ws['A1'].value == 'TEST PTA'

def test_build_treasurer_bank_reconciliation():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA')
    # Find beginning balance cell
    found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Beginning Balance':
            assert row[1] == 32630.10
            found = True
            break
    assert found, 'Beginning Balance row not found'

def test_build_treasurer_difference_is_zero():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA')
    found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'Difference (should be $0.00)':
            assert abs(row[1]) < 0.01
            found = True
            break
    assert found, 'Difference row not found'


# ── Pass-through fund section tests ─────────────────────────────────────────
# (generalized from the private notebook's "READTHON" feature — commit c348462)

MOCK_PASS_THROUGH = {
    'fund_name':      'READTHON',
    'income_total':   500.0,
    'expense_total':  300.0,
    'net':            200.0,
    'balance_held':   850.0,
}


def test_build_treasurer_without_pass_through_no_section():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA')
    labels = {row[0] for row in ws.iter_rows(values_only=True)}
    assert 'Total Money in Account' not in labels
    assert 'PTA Money' not in labels


def test_build_treasurer_pass_through_section_present():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA',
                    pass_through=MOCK_PASS_THROUGH)
    labels = {row[0] for row in ws.iter_rows(values_only=True)}
    for expected in ['(+) READTHON Deposits This Month',
                      '(-) READTHON Payouts This Month',
                      'READTHON Balance Held in Account (Cumulative)',
                      'Total Money in Account',
                      'PTA Money']:
        assert expected in labels, f'{expected!r} not found'


def test_build_treasurer_pass_through_values():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA',
                    pass_through=MOCK_PASS_THROUGH)
    values = {row[0]: row[1] for row in ws.iter_rows(values_only=True)}
    assert values['(+) READTHON Deposits This Month'] == 500.0
    assert values['(-) READTHON Payouts This Month']  == -300.0
    assert values['READTHON Balance Held in Account (Cumulative)'] == 850.0


def test_build_treasurer_pta_money_formula():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA',
                    pass_through=MOCK_PASS_THROUGH)
    pta_cell = None
    for row_cells in ws.iter_rows():
        if row_cells[0].value == 'PTA Money':
            pta_cell = row_cells[1]
            break
    assert pta_cell is not None, 'PTA Money row not found'
    assert isinstance(pta_cell.value, str) and pta_cell.value.startswith('=B')
    assert pta_cell.value.count('-B') == 1


def test_build_treasurer_pta_money_negative_flags_red():
    wb = openpyxl.Workbook()
    ws = wb.active
    large_balance = {**MOCK_PASS_THROUGH, 'balance_held': 99999.0}  # > ending_balance
    build_treasurer(ws, MOCK_QRB, MOCK_BANK, 'July 2025', 'Setauket School PTA',
                    pass_through=large_balance)
    pta_cell = None
    for row_cells in ws.iter_rows():
        if row_cells[0].value == 'PTA Money':
            pta_cell = row_cells[1]
            break
    assert pta_cell is not None, 'PTA Money row not found'
    assert pta_cell.fill == RED_FILL
    assert pta_cell.fill != GOLD_FILL


def test_build_ytd_summary_dates_reflect_fiscal_year_start(sample_merged_income, sample_merged_expense):
    # Previously hardcoded to literal "7/1/25" / "as of 6/30/25" regardless
    # of which fiscal year the report was actually for.
    wb = openpyxl.Workbook()
    ws = wb.active
    build_ytd_summary(ws, sample_merged_income, sample_merged_expense, 'Test PTA',
                       'July 2026', 0, FISCAL_MONTHS, 2026, bank=MOCK_BANK)
    assert '7/1/2026' in ws['A2'].value
    assert ws['C3'].value == 'as of 6/30/2027'


# ── YTD Summary PTA Money tests ─────────────────────────────────────────────

def test_build_ytd_summary_without_pass_through_no_pta_money(sample_merged_income, sample_merged_expense):
    wb = openpyxl.Workbook()
    ws = wb.active
    build_ytd_summary(ws, sample_merged_income, sample_merged_expense, 'Test PTA',
                       'July 2025', 0, FISCAL_MONTHS, 2025, bank=MOCK_BANK)
    assert ws['H3'].value is None
    assert ws['I3'].value is None


def test_build_ytd_summary_pass_through_zero_balance_no_pta_money(sample_merged_income, sample_merged_expense):
    wb = openpyxl.Workbook()
    ws = wb.active
    zero_balance = {**MOCK_PASS_THROUGH, 'balance_held': 0.0}
    build_ytd_summary(ws, sample_merged_income, sample_merged_expense, 'Test PTA',
                       'July 2025', 0, FISCAL_MONTHS, 2025, bank=MOCK_BANK, pass_through=zero_balance)
    # nothing to subtract, so PTA Money would just duplicate Current Balance -- hidden
    assert ws['H3'].value is None
    assert ws['I3'].value is None


def test_build_ytd_summary_pass_through_nonzero_balance_shows_pta_money(sample_merged_income, sample_merged_expense):
    wb = openpyxl.Workbook()
    ws = wb.active
    build_ytd_summary(ws, sample_merged_income, sample_merged_expense, 'Test PTA',
                       'July 2025', 0, FISCAL_MONTHS, 2025, bank=MOCK_BANK, pass_through=MOCK_PASS_THROUGH)
    assert ws['H3'].value == 'PTA Money:'
    assert ws['I3'].value == MOCK_BANK['ending_balance'] - MOCK_PASS_THROUGH['balance_held']


# ── Budget sheet tests ────────────────────────────────────────────────────────

def test_build_budget_income(sample_merged_income):
    wb = openpyxl.Workbook()
    ws = wb.active
    build_budget(ws, 'Budget vs Actuals - Income',
                 sample_merged_income, 'Setauket School PTA',
                 FISCAL_MONTHS, 0, 2025, show_pl=False)
    assert 'SETAUKET SCHOOL PTA' in ws['A1'].value

def test_build_budget_fiscal_year_label_reflects_fiscal_year_start(sample_merged_income):
    # Previously hardcoded to a literal "Fiscal Year July 2025 - June 2026"
    # string regardless of which fiscal year the report was actually for --
    # confirmed wrong in a real FY2026-27 report. Must track fiscal_year_start.
    wb = openpyxl.Workbook()
    ws = wb.active
    build_budget(ws, 'Test', sample_merged_income,
                 'Test PTA', FISCAL_MONTHS, 0, 2026, show_pl=False)
    assert 'Fiscal Year 2026 - 2027' in ws['A2'].value
    assert '2025 - 2026' not in ws['A2'].value

def test_build_budget_active_month_highlighted(sample_merged_income):
    wb = openpyxl.Workbook()
    ws = wb.active
    build_budget(ws, 'Test', sample_merged_income,
                 'Test PTA', FISCAL_MONTHS, 0, 2025, show_pl=False)
    # Column D (index 4) = JULY = fiscal index 0 — should be gold
    header_cell = ws.cell(row=3, column=4)
    # openpyxl stores colors as 8-char ARGB (alpha + RGB)
    # so FFD966 becomes 00FFD966
    assert header_cell.fill.fgColor.rgb.endswith('FFD966')  # ← use endswith

def test_build_budget_expense_with_pl(sample_merged_expense, sample_merged_income):
    wb = openpyxl.Workbook()
    ws = wb.active
    build_budget(ws, 'Budget vs Actuals - Expenses',
                 sample_merged_expense, 'Test PTA',
                 FISCAL_MONTHS, 0, 2025, show_pl=True,
                 income_merged=sample_merged_income)
    # Should have Profit/Loss header
    found = False
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        if 'Profit/Loss' in [v for v in row if v]:
            found = True
    assert found, 'Profit/Loss column not found'


# ── Givebacks reconciliation tests ───────────────────────────────────────────

def test_build_givebacks_total():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_givebacks(ws, MOCK_GIVEBACKS, MOCK_BANK, 'Test PTA')
    # Check total row exists
    found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'TOTAL':
            found = True
            break
    assert found, 'TOTAL row not found'

def test_build_givebacks_item_count():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_givebacks(ws, MOCK_GIVEBACKS, MOCK_BANK, 'Test PTA')
    items_found = 0
    for row in ws.iter_rows(values_only=True):
        if row[0] in ('Shop to Give Donation', 'Teacher/Staff'):
            items_found += 1
    assert items_found == 2


# ── Fiscal months constants test ──────────────────────────────────────────────

def test_fiscal_months_count():
    assert len(FISCAL_MONTHS) == 12

def test_fiscal_months_order():
    assert FISCAL_MONTHS[0]  == 'JULY'
    assert FISCAL_MONTHS[5]  == 'DEC'
    assert FISCAL_MONTHS[6]  == 'JAN'
    assert FISCAL_MONTHS[11] == 'JUNE'


# ── Debits & Credits (whole-fiscal-year ledger) tests ─────────────────────────
# (ported from the private notebook's Debits & Credits ledger work)

MOCK_CREDITS_BY_MONTH = [
    ('July 2025', [
        {'date': '07/17/2025', 'type': 'Deposit', 'check_no': '', 'payee': 'MemberHub',
         'description': 'Deposit', 'category': 'Membership', 'amount': 110.14, 'is_income': True,
         'bank_statement_month': 'August 2025'},  # lag case
    ]),
    ('August 2025', [
        {'date': '08/05/2025', 'type': 'Deposit', 'check_no': '', 'payee': 'MemberHub',
         'description': 'Deposit', 'category': 'Book Fair', 'amount': 50.0, 'is_income': True,
         'bank_statement_month': 'August 2025'},  # same-month match
    ]),
    ('September 2025', []),  # no credits that month - should be skipped, not error
]

MOCK_QB_TO_BUDGET_MAP = {'Membership': 'Membership Income', 'Picnic': 'Picnic Fund'}


def _cell_values(ws):
    return [row for row in ws.iter_rows(values_only=True) if any(v is not None for v in row)]


def test_build_credits_sheet_running_total():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, MOCK_CREDITS_BY_MONTH, 'Test PTA', MOCK_QB_TO_BUDGET_MAP)
    rows = _cell_values(ws)
    amounts = [r[6] for r in rows if isinstance(r[0], str) and '/' in str(r[0])]
    assert amounts == [110.14, 160.14]  # cumulative across months


def test_build_credits_sheet_skips_empty_month():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, MOCK_CREDITS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    month_bands = [r[0] for r in rows if r[0] in ('JULY', 'AUGUST', 'SEPTEMBER')]
    assert month_bands == ['JULY', 'AUGUST']  # September (empty) skipped


def test_build_credits_sheet_budget_line_mapping():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, MOCK_CREDITS_BY_MONTH, 'Test PTA', MOCK_QB_TO_BUDGET_MAP)
    rows = _cell_values(ws)
    budget_lines = [r[5] for r in rows if isinstance(r[0], str) and '/' in str(r[0])]
    assert budget_lines == ['Membership Income', 'Book Fair']  # mapped / falls back to raw category


def test_build_credits_sheet_total_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, MOCK_CREDITS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    total_rows = [r for r in rows if r[0] == 'TOTAL CREDITS']
    assert len(total_rows) == 1
    assert total_rows[0][2] == 160.14


def test_build_credits_sheet_bank_statement_column():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, MOCK_CREDITS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    header = [r for r in rows if r[0] == 'DEPOSIT DATE'][0]
    assert header[4] == 'BANK STATEMENT'
    bank_stmts = [r[4] for r in rows if isinstance(r[0], str) and '/' in str(r[0])]
    assert bank_stmts == ['August', 'August']  # July txn lagged into August's statement


def test_build_credits_sheet_labels_raw_bank_deposits():
    credits_by_month = [
        ('September 2025', [
            {'date': '09/24/2025', 'type': 'Deposit', 'check_no': '', 'payee': '',
             'description': 'DEPOSIT ID NUMBER XX8881', 'category': 'Book Fair',
             'amount': 4365.50, 'is_income': True, 'bank_statement_month': 'September 2025'},
            {'date': '09/12/2025', 'type': 'Deposit', 'check_no': '', 'payee': '',
             'description': 'DEPOSIT', 'category': 'Holiday Boutique',
             'amount': 500.0, 'is_income': True, 'bank_statement_month': 'September 2025'},
            {'date': '09/19/2025', 'type': 'Deposit', 'check_no': '', 'payee': '',
             'description': 'MemberHub/Givebacks Deposit', 'category': 'MemberHub/Givebacks Deposit',
             'amount': 90.0, 'is_income': True, 'bank_statement_month': 'September 2025'},
        ]),
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, credits_by_month, 'Test PTA',
                         {'Book Fair': 'Book Fair', 'Holiday Boutique': 'Holiday Boutique'})
    rows = _cell_values(ws)
    data_rows = {r[0]: r for r in rows if isinstance(r[0], str) and '/' in str(r[0])}
    # both raw cash/check deposits show the generic category label...
    assert data_rows['09/24/2025'][1] == 'Bank Deposit (cash/check)'
    assert data_rows['09/12/2025'][1] == 'Bank Deposit (cash/check)'
    # ...but keep their real budget line
    assert data_rows['09/24/2025'][5] == 'Book Fair'
    assert data_rows['09/12/2025'][5] == 'Holiday Boutique'
    # a Givebacks-sourced row is untouched
    assert data_rows['09/19/2025'][1] == 'MemberHub/Givebacks Deposit'


def test_build_credits_sheet_nests_multi_category_bank_deposit():
    # Real scenario: one $7,118.12 Chase deposit that QuickBooks split
    # across two categories - should render as one 'Bank Deposit' band
    # (the real, auditable total) with Book Fair/Spiritwear nested beneath,
    # not two unrelated-looking flat rows.
    credits_by_month = [
        ('February 2026', [
            {'date': '02/27/2026', 'type': 'Deposit', 'check_no': '', 'payee': '',
             'description': 'DEPOSIT', 'category': 'Book Fair', 'amount': 6219.40,
             'is_income': True, 'bank_statement_month': 'February 2026'},
            {'date': '02/27/2026', 'type': 'Deposit', 'check_no': '', 'payee': '',
             'description': 'DEPOSIT', 'category': 'Spiritwear', 'amount': 898.72,
             'is_income': True, 'bank_statement_month': 'February 2026'},
        ]),
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, credits_by_month, 'Test PTA',
                         {'Book Fair': 'Book Fair', 'Spiritwear': 'Spiritwear'})
    rows = _cell_values(ws)

    band_rows = [r for r in rows if r[1] == 'Bank Deposit']
    assert len(band_rows) == 1
    assert band_rows[0][0] == '02/27/2026'
    assert round(band_rows[0][2], 2) == 7118.12

    categories = [r[1] for r in rows if r[1] in ('Book Fair', 'Spiritwear')]
    assert sorted(categories) == ['Book Fair', 'Spiritwear']
    amounts = sorted(r[2] for r in rows if r[1] in ('Book Fair', 'Spiritwear'))
    assert amounts == [898.72, 6219.40]
    # running total still progresses per underlying category row
    running_totals = sorted(r[6] for r in rows if r[1] in ('Book Fair', 'Spiritwear'))
    assert running_totals == [898.72, 7118.12] or running_totals == [6219.40, 7118.12]


def test_build_credits_sheet_no_band_for_unmatched_multi_row_date():
    # Same date, more than one row, but bank_statement_month is None -
    # never verified against a real deposit, so must NOT be grouped into a
    # band (that would be claiming a reconciliation that doesn't exist).
    credits_by_month = [
        ('February 2026', [
            {'date': '02/27/2026', 'type': 'Deposit', 'check_no': '', 'payee': '',
             'description': 'DEPOSIT', 'category': 'Book Fair', 'amount': 100.0,
             'is_income': True, 'bank_statement_month': None},
            {'date': '02/27/2026', 'type': 'Deposit', 'check_no': '', 'payee': '',
             'description': 'DEPOSIT', 'category': 'Spiritwear', 'amount': 50.0,
             'is_income': True, 'bank_statement_month': None},
        ]),
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    build_credits_sheet(ws, credits_by_month, 'Test PTA')
    rows = _cell_values(ws)
    assert not any(r[1] == 'Bank Deposit' for r in rows)


MOCK_DEBITS_BY_MONTH = [
    ('July 2025', [
        {'date': '07/01/2025', 'type': 'Check', 'check_no': '1077', 'payee': 'Jane Doe',
         'description': 'CHECK # 1077', 'category': 'Picnic', 'amount': 181.58, 'is_income': False,
         'bank_statement_month': 'July 2025'},
        {'date': '07/17/2025', 'type': 'Expense', 'check_no': '', 'payee': 'Quickbooks Online',
         'description': 'Accounting', 'category': 'Accounting Expense (Quickbooks)',
         'amount': 1257.76, 'is_income': False,
         'bank_statement_month': None},  # unmatched - should render as '—'
    ]),
]

MOCK_GIVEBACKS_BY_MONTH = [
    ('July 2025', [
        {'date': '07/10/2025', 'total': 110.14, 'items': [
            {'item': 'Shop to Give Donation', 'category': '', 'count': 1, 'total': 95.14, 'source_file': 'x.csv'},
            {'item': 'Teacher/Staff', 'category': 'Memberships', 'count': 1, 'total': 15.0, 'source_file': 'x.csv'},
        ]},
        {'date': None, 'total': 50.0, 'items': [
            {'item': 'Family Membership', 'category': '', 'count': 1, 'total': 50.0, 'source_file': 'y.csv'},
        ]},
    ]),
]


def test_build_debits_sheet_running_total_and_notes_blank():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_debits_sheet(ws, MOCK_DEBITS_BY_MONTH, 'Test PTA', MOCK_QB_TO_BUDGET_MAP)
    rows = _cell_values(ws)
    data_rows = [r for r in rows if r[0] in ('1077', '')]
    assert len(data_rows) == 2
    assert data_rows[1][9] == 1439.34  # running total after both debits
    for r in data_rows:
        assert not r[8]  # NOTES column intentionally blank


def test_build_debits_sheet_bank_statement_column():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_debits_sheet(ws, MOCK_DEBITS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    header = [r for r in rows if r[0] == 'CHECK #'][0]
    assert header[6] == 'BANK STATEMENT'
    data_rows = [r for r in rows if r[0] in ('1077', '')]
    assert data_rows[0][6] == 'July'    # matched
    assert data_rows[1][6] == '—'       # unmatched (bank_statement_month=None)


def test_build_debits_sheet_total_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_debits_sheet(ws, MOCK_DEBITS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    total_rows = [r for r in rows if r[0] == 'TOTAL DEBITS']
    assert len(total_rows) == 1
    assert total_rows[0][4] == 1439.34


def test_build_memberhub_summary_payout_bands_present():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_memberhub_summary_sheet(ws, MOCK_GIVEBACKS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    payout_labels = [r[0] for r in rows if isinstance(r[0], str) and r[0].startswith('Payout')]
    # dated payout sorts before the unreconciled (date=None) one
    assert payout_labels == ['Payout — 07/10/2025', 'Payout — (date not reconciled)']


def test_build_memberhub_summary_payout_band_shows_total():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_memberhub_summary_sheet(ws, MOCK_GIVEBACKS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    dated_payout = [r for r in rows if r[0] == 'Payout — 07/10/2025'][0]
    assert dated_payout[2] == 110.14


def test_build_memberhub_summary_items_nested_under_payout():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_memberhub_summary_sheet(ws, MOCK_GIVEBACKS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    item_rows = [r for r in rows
                if r[0] in ('Shop to Give Donation', 'Teacher/Staff', 'Family Membership')]
    assert len(item_rows) == 3
    assert item_rows[0][2] == 95.14
    assert item_rows[2][2] == 50.0  # unreconciled payout's item still rendered, not dropped


def test_build_memberhub_summary_running_total():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_memberhub_summary_sheet(ws, MOCK_GIVEBACKS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    item_rows = [r for r in rows
                if r[0] in ('Shop to Give Donation', 'Teacher/Staff', 'Family Membership')]
    assert [r[3] for r in item_rows] == [95.14, 110.14, 160.14]


def test_build_memberhub_summary_total_row():
    wb = openpyxl.Workbook()
    ws = wb.active
    build_memberhub_summary_sheet(ws, MOCK_GIVEBACKS_BY_MONTH, 'Test PTA')
    rows = _cell_values(ws)
    total_rows = [r for r in rows if r[0] == 'TOTAL GIVEBACKS']
    assert len(total_rows) == 1
    assert total_rows[0][2] == 160.14
