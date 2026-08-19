"""
Tests for pipeline.py — runs the full pipeline against the synthetic
sample_data/July_1999/ dataset.
"""
import json
import shutil
from datetime import date
from pathlib import Path

import openpyxl
import pytest

from pta_treasurer.budget_io import generate_template
from pta_treasurer.config import OrgConfig, fiscal_year_start_calendar_year
from pta_treasurer.pipeline import (
    _apply_oct_2025_tshirt_split, build_debits_credits_ledger,
    compute_pass_through_balance_held, get_fiscal_year_balance_forward,
    list_available_fiscal_years, load_all_actuals, run_month,
)

SAMPLE_DATA = Path(__file__).parent.parent / 'sample_data' / 'July_1999'


@pytest.fixture
def data_dir(tmp_path):
    """A data folder with sample_data/July_1999 as input, plus a budget
    template matching the QuickBooks categories in that sample month."""
    month_folder = tmp_path / 'input' / 'July_1999'
    shutil.copytree(SAMPLE_DATA, month_folder)

    budget_path = tmp_path / 'budget.xlsx'
    generate_template(budget_path)
    wb = openpyxl.load_workbook(budget_path)
    wb['Income Budget'].append(['Fundraising', 'Book Fair', '', 0, 500.0])
    ws = wb['Expense Budget']
    ws.append(['Admin/General', 'Accounting Expense (Quickbooks)', '', 0, 1300.0])
    ws.append(['Program Expense', 'Picnic', '', 0, 200.0])
    wb.save(budget_path)

    return tmp_path


@pytest.fixture
def config():
    return OrgConfig(org_name='Demo School PTA', balance_forward=0.0)


def test_run_month_produces_workbook_with_six_tabs(data_dir, config):
    result = run_month(config, data_dir, 'July', '1999')
    assert result.output_path.exists()

    wb = openpyxl.load_workbook(result.output_path)
    assert wb.sheetnames == [
        'Treasurer Report', 'YTD Summary', 'Income Budget vs Actuals',
        'Expense Budget vs Actuals', 'Giveback Reconciliation', 'File Manifest',
    ]


def test_run_month_parses_correct_totals(data_dir, config):
    result = run_month(config, data_dir, 'July', '1999')
    assert result.qb['expense_total'] == pytest.approx(1439.34)
    assert result.bank['beginning_balance'] == pytest.approx(32630.10)
    assert result.bank['ending_balance'] == pytest.approx(31190.76)
    assert len(result.givebacks) == 2
    assert result.warnings == []


def test_run_month_writes_history_entry(data_dir, config):
    run_month(config, data_dir, 'July', '1999')
    hist_path = data_dir / 'data' / 'history' / '1999_to_2000' / 'July_1999.json'
    assert hist_path.exists()
    entry = json.loads(hist_path.read_text())
    assert entry['month_label'] == 'July 1999'
    assert entry['fiscal_index'] == 0
    assert entry['expense_total'] == pytest.approx(1439.34)


def test_run_month_output_filename_includes_fiscal_year(data_dir, config):
    result = run_month(config, data_dir, 'July', '1999')
    assert result.output_path.name == 'Treasurer_Report_1999_to_2000_July_1999.xlsx'


def test_run_month_merges_history_into_budget(data_dir, config):
    run_month(config, data_dir, 'July', '1999')
    wb = openpyxl.load_workbook(
        data_dir / 'output' / 'Treasurer_Report_1999_to_2000_July_1999.xlsx'
    )
    ws = wb['Expense Budget vs Actuals']
    # Column D (index 4) is JULY (fiscal index 0); find the Picnic row.
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] == 'Picnic':
            assert row[3] == pytest.approx(181.58)
            break
    else:
        pytest.fail('Picnic row not found in Expense Budget vs Actuals')


def test_run_month_without_givebacks_folder_warns_but_succeeds(data_dir, config):
    shutil.rmtree(data_dir / 'input' / 'July_1999' / 'givebacks')
    result = run_month(config, data_dir, 'July', '1999')
    assert result.givebacks == []
    assert any('Givebacks' in w for w in result.warnings)
    assert result.output_path.exists()


def test_run_month_missing_input_folder_raises(tmp_path, config):
    budget_path = tmp_path / 'budget.xlsx'
    generate_template(budget_path)
    with pytest.raises(FileNotFoundError):
        run_month(config, tmp_path, 'July', '1999')


def test_run_month_missing_budget_template_raises(data_dir, config):
    (data_dir / 'budget.xlsx').unlink()
    with pytest.raises(FileNotFoundError):
        run_month(config, data_dir, 'July', '1999')


def test_load_all_actuals_empty_when_no_history(tmp_path):
    income, expense = load_all_actuals(tmp_path, fiscal_year_start=1999)
    assert income == {}
    assert expense == {}


def test_load_all_actuals_partitioned_storage_keeps_fiscal_years_separate(tmp_path):
    # History is partitioned on disk per fiscal year specifically so that
    # two different fiscal years' Julys (both fiscal_index 0) can never
    # collide -- this is what a real production bug looked like before the
    # storage was restructured: both years' data lived in one flat
    # directory and the alphabetically-later file silently won.
    history_base = tmp_path / 'data' / 'history'
    (history_base / '2025_to_2026').mkdir(parents=True)
    (history_base / '2025_to_2026' / 'July_2025.json').write_text(json.dumps({
        'month_label': 'July 2025', 'fiscal_index': 0,
        'income': {}, 'expenses': {'Accounting Quickbooks': 1257.76},
    }))
    (history_base / '2026_to_2027').mkdir(parents=True)
    (history_base / '2026_to_2027' / 'July_2026.json').write_text(json.dumps({
        'month_label': 'July 2026', 'fiscal_index': 0,
        'income': {}, 'expenses': {'Accounting Quickbooks': 1350.68},
    }))

    income_25, expense_25 = load_all_actuals(tmp_path, fiscal_year_start=2025)
    assert expense_25['Accounting Quickbooks'][0] == 1257.76

    income_26, expense_26 = load_all_actuals(tmp_path, fiscal_year_start=2026)
    assert expense_26['Accounting Quickbooks'][0] == 1350.68


def test_run_month_uses_only_its_own_fiscal_year_actuals(data_dir, config):
    # A history entry from a DIFFERENT fiscal year's July exists in its own
    # subfolder alongside the current one. run_month() must not let it
    # bleed into this fiscal year's report.
    other_fy_history = data_dir / 'data' / 'history' / '1998_to_1999' / 'July_1998.json'
    other_fy_history.parent.mkdir(parents=True, exist_ok=True)
    other_fy_history.write_text(json.dumps({
        'month_label': 'July 1998', 'fiscal_index': 0,
        'income': {}, 'expenses': {'Accounting Expense (Quickbooks)': 999999.99},
    }))
    result = run_month(config, data_dir, 'July', '1999')
    # July 1999's own real actual, not the unrelated fiscal year's number
    assert result.expense_merged['Admin/General']['Accounting Expense (Quickbooks)'][2][0] == 1257.76


def test_run_month_uses_real_prior_fiscal_year_actual_as_last_year(data_dir, config):
    # budget.xlsx's static 'Last Year Actual' for Picnic is 0 (see the
    # data_dir fixture). A real history entry from the PRIOR fiscal year
    # (1998_to_1999, since July 1999 belongs to FY1999-2000) should override
    # that stale static value.
    prior_history = data_dir / 'data' / 'history' / '1998_to_1999' / 'March_1999.json'
    prior_history.parent.mkdir(parents=True, exist_ok=True)
    prior_history.write_text(json.dumps({
        'month_label': 'March 1999', 'fiscal_index': 8,
        'income': {}, 'expenses': {'Picnic': 999.99},
    }))
    result = run_month(config, data_dir, 'July', '1999')
    assert result.expense_merged['Program Expense']['Picnic'][0] == pytest.approx(999.99)


def test_run_month_no_prior_fiscal_year_history_keeps_static_last_year(data_dir, config):
    # No history exists for FY1998-99 at all -- budget.xlsx's static value
    # (0.0, from the data_dir fixture) is the only value there is.
    result = run_month(config, data_dir, 'July', '1999')
    assert result.expense_merged['Program Expense']['Picnic'][0] == 0.0


def test_run_month_july_2026_folds_real_actuals_into_fy2025_26_last_year(tmp_path, config):
    # One-time correction: July-2026 QuickBooks expenses are actually late
    # bills for FY2025-26 activities, so (except Accounting Quickbooks --
    # July 2026's own recurring bookkeeping fee, not a late bill) they
    # should count toward FY2025-26's ("last year's") total when
    # generating a FY2026-27 report. Pulled dynamically from July 2026's
    # REAL recorded actuals, not a frozen dollar-amount snapshot -- even
    # though no other digital history exists for FY2025-26.
    month_folder = tmp_path / 'input' / 'July_2026'
    shutil.copytree(SAMPLE_DATA, month_folder)

    budget_path = tmp_path / 'budget.xlsx'
    generate_template(budget_path)
    wb = openpyxl.load_workbook(budget_path)
    ws = wb['Expense Budget']
    ws.append(['Program Expense', 'Picnic', '', 0, 200.0])
    ws.append(['Admin/General', 'Accounting Quickbooks',
               'Accounting Expense (Quickbooks)', 0, 1300.0])
    wb.save(budget_path)

    result = run_month(config, tmp_path, 'July', '2026')
    expense = result.expense_merged
    # Picnic's real July 2026 actual (181.58, from sample_data) folds into
    # FY2025-26's last-year total.
    assert expense['Program Expense']['Picnic'][0] == pytest.approx(181.58)
    # Accounting Quickbooks is explicitly excluded from the spillover.
    assert expense['Admin/General']['Accounting Quickbooks'][0] == 0.0


def test_run_month_july_2026_spillover_zeros_items_with_no_july_actual(tmp_path, config):
    # There's no other FY2025-26 history in this tmp_path, but the July
    # 2026 spillover itself makes prior_actuals_mapped non-empty -- so an
    # item with no matching actual at all (from spillover or otherwise)
    # becomes a genuine 0.0, same as apply_dynamic_last_year's normal rule
    # once ANY prior-year data exists. Not expected to bite in production,
    # where a fiscal year with a July spillover already has its own full
    # digital history recorded well before the correction matters.
    month_folder = tmp_path / 'input' / 'July_2026'
    shutil.copytree(SAMPLE_DATA, month_folder)

    budget_path = tmp_path / 'budget.xlsx'
    generate_template(budget_path)
    wb = openpyxl.load_workbook(budget_path)
    ws = wb['Expense Budget']
    ws.append(['Program Expense', 'Never Spent On', '', 500.0, 200.0])
    wb.save(budget_path)

    result = run_month(config, tmp_path, 'July', '2026')
    assert result.expense_merged['Program Expense']['Never Spent On'][0] == 0.0


def test_run_month_spillover_does_not_apply_to_other_fiscal_years(data_dir, config):
    # The spillover correction is guarded to only fire when computing
    # FY2025-26 specifically as the prior fiscal year -- July 1999's prior
    # fiscal year is 1998-99, so the correction must stay inert.
    result = run_month(config, data_dir, 'July', '1999')
    expense = result.expense_merged.get('Program Expense', {})
    assert expense.get('Picnic', (0.0,))[0] == 0.0


# ── Tests for _apply_oct_2025_tshirt_split ──────────────────────────────────

def test_apply_oct_2025_tshirt_split_moves_125_to_staff_line():
    actuals = {'5th grade T-shirts': [0.0] * 3 + [901.0] + [0.0] * 8}
    result = _apply_oct_2025_tshirt_split(actuals, fiscal_year_start=2025)
    assert result['5th grade T-shirts'][3] == pytest.approx(776.0)
    assert result['5th Staff T-Shirts'][3] == pytest.approx(125.0)


def test_apply_oct_2025_tshirt_split_adds_to_existing_staff_actual():
    actuals = {
        '5th grade T-shirts': [0.0] * 3 + [901.0] + [0.0] * 8,
        '5th Staff T-Shirts': [0.0] * 3 + [66.78] + [0.0] * 8,
    }
    result = _apply_oct_2025_tshirt_split(actuals, fiscal_year_start=2025)
    assert result['5th Staff T-Shirts'][3] == pytest.approx(66.78 + 125.0)


def test_apply_oct_2025_tshirt_split_inert_for_other_fiscal_years():
    actuals = {'5th grade T-shirts': [0.0] * 3 + [901.0] + [0.0] * 8}
    result = _apply_oct_2025_tshirt_split(actuals, fiscal_year_start=2026)
    assert result == actuals


def test_apply_oct_2025_tshirt_split_noop_when_no_october_actual():
    actuals = {'5th grade T-shirts': [0.0] * 12}
    result = _apply_oct_2025_tshirt_split(actuals, fiscal_year_start=2025)
    assert result == actuals


def test_apply_oct_2025_tshirt_split_noop_when_category_absent():
    actuals = {'Some Other Category': [10.0] * 12}
    result = _apply_oct_2025_tshirt_split(actuals, fiscal_year_start=2025)
    assert result == actuals


# ── Tests for get_fiscal_year_balance_forward ───────────────────────────────

def test_get_fiscal_year_balance_forward_reuses_bank_when_processing_july(tmp_path, config):
    bank = {'beginning_balance': 500.0}
    result, warning = get_fiscal_year_balance_forward(
        config, tmp_path, 1999, 'July 1999', bank)
    assert result == 500.0
    assert warning is None


def test_get_fiscal_year_balance_forward_reads_july_statement_for_later_month(data_dir, config):
    # data_dir's fixture already has a real July_1999 folder (sample_data)
    # whose Chase PDF beginning_balance is 32630.10 -- must be read from
    # disk, not from the CURRENT month's own (unrelated) bank object.
    unrelated_bank = {'beginning_balance': 999999.99}
    result, warning = get_fiscal_year_balance_forward(
        config, data_dir, 1999, 'August 1999', unrelated_bank)
    assert result == pytest.approx(32630.10)
    assert warning is None


def test_get_fiscal_year_balance_forward_subtracts_pass_through_money(data_dir):
    # Balance Forward should represent the org's OWN money, not
    # pass-through fund money (e.g. READTHON) already sitting in the
    # account at that fiscal year's July 1 -- same "PTA Money" concept
    # already used for Current Balance.
    config = OrgConfig(org_name='Demo School PTA',
                        pass_through_fund_name='READTHON',
                        pass_through_fund_categories='READTHON')
    other_fy_history = data_dir / 'data' / 'history' / '1998_to_1999' / 'June_1999.json'
    other_fy_history.parent.mkdir(parents=True, exist_ok=True)
    other_fy_history.write_text(json.dumps({
        'month_label': 'June 1999', 'pass_through_net': 300.0,
    }))
    unrelated_bank = {'beginning_balance': 32630.10}
    result, warning = get_fiscal_year_balance_forward(
        config, data_dir, 1999, 'July 1999', unrelated_bank)
    assert result == pytest.approx(32630.10 - 300.0)
    assert warning is None


def test_get_fiscal_year_balance_forward_no_pass_through_configured_keeps_raw_balance(data_dir, config):
    # config (the fixture) has no pass_through_fund_name set -- must not
    # subtract anything even if unrelated pass-through-shaped history exists.
    bank = {'beginning_balance': 32630.10}
    result, warning = get_fiscal_year_balance_forward(
        config, data_dir, 1999, 'July 1999', bank)
    assert result == pytest.approx(32630.10)


def test_get_fiscal_year_balance_forward_no_july_folder_falls_back_to_static(tmp_path):
    config = OrgConfig(org_name='Demo School PTA', balance_forward=1234.56)
    result, warning = get_fiscal_year_balance_forward(
        config, tmp_path, 1999, 'August 1999', {'beginning_balance': 0.0})
    assert result == 1234.56
    assert warning is None


def test_get_fiscal_year_balance_forward_unparseable_pdf_falls_back_with_warning(tmp_path):
    config = OrgConfig(org_name='Demo School PTA', balance_forward=1234.56)
    july_folder = tmp_path / 'input' / 'July_1999'
    july_folder.mkdir(parents=True)
    (july_folder / 'not_a_real_statement.pdf').write_text('not a pdf')
    result, warning = get_fiscal_year_balance_forward(
        config, tmp_path, 1999, 'August 1999', {'beginning_balance': 0.0})
    assert result == 1234.56
    assert warning is not None


def test_run_month_ytd_summary_uses_dynamic_balance_forward(data_dir, config):
    # config.balance_forward defaults to 0.0 in the `config` fixture -- the
    # YTD Summary's Balance Forward must come from July 1999's own bank
    # statement instead, not the stale/unset static value.
    result = run_month(config, data_dir, 'July', '1999')
    wb = openpyxl.load_workbook(result.output_path)
    ws = wb['YTD Summary']
    assert ws['B3'].value == pytest.approx(32630.10)


def test_compute_pass_through_balance_held_no_history_returns_forward(tmp_path):
    assert compute_pass_through_balance_held(tmp_path, 100.0) == 100.0


def test_compute_pass_through_balance_held_sums_in_true_calendar_order(tmp_path):
    # Written deliberately out of calendar order (June before July, and a
    # January entry) -- filename/glob order would sum these wrong; only
    # true calendar-date sorting gets it right. All three fall within
    # FY2025-26, so they share one fiscal-year subfolder.
    history_dir = tmp_path / 'data' / 'history' / '2025_to_2026'
    history_dir.mkdir(parents=True)
    entries = {
        'June_2026.json':    {'month_label': 'June 2026', 'pass_through_net': 50.0},
        'July_2025.json':    {'month_label': 'July 2025', 'pass_through_net': 200.0},
        'January_2026.json': {'month_label': 'January 2026', 'pass_through_net': -30.0},
    }
    for fname, entry in entries.items():
        (history_dir / fname).write_text(json.dumps(entry))

    balance = compute_pass_through_balance_held(tmp_path, balance_forward=100.0)
    assert balance == pytest.approx(100.0 + 200.0 - 30.0 + 50.0)


def test_compute_pass_through_balance_held_sums_across_fiscal_year_subfolders(tmp_path):
    # Pass-through fund money (e.g. a fundraiser's proceeds) carries over
    # indefinitely and isn't the org's own money, so its running balance
    # must span fiscal-year boundaries -- unlike load_all_actuals(), this
    # must aggregate across MULTIPLE fiscal-year subfolders, not just one.
    history_base = tmp_path / 'data' / 'history'
    (history_base / '2024_to_2025').mkdir(parents=True)
    (history_base / '2024_to_2025' / 'June_2025.json').write_text(json.dumps({
        'month_label': 'June 2025', 'pass_through_net': 300.0,
    }))
    (history_base / '2025_to_2026').mkdir(parents=True)
    (history_base / '2025_to_2026' / 'July_2025.json').write_text(json.dumps({
        'month_label': 'July 2025', 'pass_through_net': 200.0,
    }))

    balance = compute_pass_through_balance_held(tmp_path, balance_forward=0.0)
    assert balance == pytest.approx(300.0 + 200.0)


def test_compute_pass_through_balance_held_before_fiscal_year_start_excludes_that_year(tmp_path):
    # Used by get_fiscal_year_balance_forward() to get the balance held as
    # of a fiscal year's own July 1 -- must include everything from EARLIER
    # fiscal years but exclude that fiscal year's own entries (including
    # its own July, dated exactly July 1).
    history_base = tmp_path / 'data' / 'history'
    (history_base / '2024_to_2025').mkdir(parents=True)
    (history_base / '2024_to_2025' / 'June_2025.json').write_text(json.dumps({
        'month_label': 'June 2025', 'pass_through_net': 300.0,
    }))
    (history_base / '2025_to_2026').mkdir(parents=True)
    (history_base / '2025_to_2026' / 'July_2025.json').write_text(json.dumps({
        'month_label': 'July 2025', 'pass_through_net': 200.0,
    }))

    balance = compute_pass_through_balance_held(
        tmp_path, balance_forward=0.0, before_fiscal_year_start=2025)
    assert balance == pytest.approx(300.0)  # only the 2024-25 entry


def test_compute_pass_through_balance_held_ignores_entries_missing_the_key(tmp_path):
    history_dir = tmp_path / 'data' / 'history' / '2025_to_2026'
    history_dir.mkdir(parents=True)
    (history_dir / 'July_2025.json').write_text(
        json.dumps({'month_label': 'July 2025', 'income_total': 500.0}))  # pre-feature entry
    assert compute_pass_through_balance_held(tmp_path, 0.0) == 0.0


def test_run_month_with_pass_through_fund_configured(data_dir):
    config = OrgConfig(org_name='Demo School PTA', pass_through_fund_name='READTHON',
                        pass_through_fund_categories='READTHON')
    month_folder = data_dir / 'input' / 'July_1999'
    qb_file = next(month_folder.glob('quickbooks*.csv'))
    content = qb_file.read_text()
    # Column offset matches this file's own header row (one leading empty
    # column, not the two-leading-empty shape used elsewhere in tests/conftest.py).
    content += (
        'Readthon,,,,,,,,\n'
        ',07/10/1999,Deposit,,Family A,Readathon pledge,Checking,500.00,500.00\n'
        'Total for Readthon,,,,,,,$500.00,\n'
    )
    qb_file.write_text(content)

    result = run_month(config, data_dir, 'July', '1999')
    assert result.qb['pass_through_income_total'] == 500.0
    wb = openpyxl.load_workbook(result.output_path)
    ws = wb['Treasurer Report']
    labels = {row[0] for row in ws.iter_rows(values_only=True)}
    assert 'Total Money in Account' in labels
    assert 'PTA Money' in labels

    entry = json.loads(
        (data_dir / 'data' / 'history' / '1999_to_2000' / 'July_1999.json').read_text())
    assert entry['pass_through_net'] == 500.0


# ── Tests for build_debits_credits_ledger ───────────────────────────────────
# (ported orchestration from the private notebook's Debits & Credits ledger)

@pytest.fixture
def ledger_data_dir(tmp_path):
    """A data folder with sample_data/July_1999 placed under a folder named
    for July of the CURRENT fiscal year (not literally 1999) -- the ledger
    scopes to the fiscal year containing today's date, and July is always
    that fiscal year's start month, so this keeps the test valid regardless
    of what "today" actually is when it runs. The sample files' own content
    (dates, filenames) still say 1999 internally, which is fine -- this is
    a wiring/smoke test, not a numeric-accuracy test (that's covered by the
    unit tests for the underlying parser/builder functions)."""
    today_fy_start = fiscal_year_start_calendar_year(
        f'{date.today().strftime("%B")} {date.today().year}')
    month_folder = tmp_path / 'input' / f'July_{today_fy_start}'
    shutil.copytree(SAMPLE_DATA, month_folder)

    budget_path = tmp_path / 'budget.xlsx'
    generate_template(budget_path)

    return tmp_path


def test_build_debits_credits_ledger_produces_three_sheets(ledger_data_dir, config):
    result = build_debits_credits_ledger(config, ledger_data_dir)
    assert result.output_path.exists()
    wb = openpyxl.load_workbook(result.output_path)
    assert wb.sheetnames == ['Credits', 'Debits', 'MemberHub_Summary']


def test_build_debits_credits_ledger_output_filename_matches_fiscal_year(ledger_data_dir, config):
    result = build_debits_credits_ledger(config, ledger_data_dir)
    today_fy_start = fiscal_year_start_calendar_year(
        f'{date.today().strftime("%B")} {date.today().year}')
    expected_name = f'Debits_and_Credits_{today_fy_start}_to_{today_fy_start + 1}.xlsx'
    assert result.output_path.name == expected_name


def test_build_debits_credits_ledger_no_month_folders_raises(tmp_path, config):
    with pytest.raises(FileNotFoundError):
        build_debits_credits_ledger(config, tmp_path)


def test_build_debits_credits_ledger_ignores_months_outside_current_fiscal_year(ledger_data_dir, config):
    # Add a second month folder from an unmistakably different fiscal year --
    # must not appear in the built ledger's sheets.
    other_folder = ledger_data_dir / 'input' / 'July_1901'
    shutil.copytree(SAMPLE_DATA, other_folder)

    result = build_debits_credits_ledger(config, ledger_data_dir)
    wb = openpyxl.load_workbook(result.output_path)
    ws = wb['Credits']
    labels = {row[0] for row in ws.iter_rows(values_only=True) if row[0]}
    assert '1901' not in ' '.join(str(label) for label in labels)


def test_list_available_fiscal_years_returns_most_recent_first(ledger_data_dir, config):
    other_folder = ledger_data_dir / 'input' / 'July_1901'
    shutil.copytree(SAMPLE_DATA, other_folder)

    today_fy_start = fiscal_year_start_calendar_year(
        f'{date.today().strftime("%B")} {date.today().year}')
    years = list_available_fiscal_years(ledger_data_dir)
    assert years == [today_fy_start, 1901]


def test_build_debits_credits_ledger_explicit_fiscal_year_overrides_today(ledger_data_dir, config):
    # A month folder from a fiscal year that is NOT "today's" -- only
    # reachable by explicitly requesting it, proving the override works
    # independent of the current real-world date.
    other_folder = ledger_data_dir / 'input' / 'July_1901'
    shutil.copytree(SAMPLE_DATA, other_folder)

    result = build_debits_credits_ledger(config, ledger_data_dir, fiscal_year_start=1901)
    assert result.output_path.name == 'Debits_and_Credits_1901_to_1902.xlsx'
