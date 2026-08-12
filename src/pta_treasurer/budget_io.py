"""
budget_io.py
Reads and writes the Excel-template budget config that replaces the old
notebook's hardcoded INCOME_BUDGET / EXPENSE_BUDGET / QB_TO_BUDGET_MAP dicts.

The functions below `merge_actuals_into_budget` (add_item/remove_item/
move_item/rename_item/map_qb_category/set_budget_amount, dispatched via
apply_edit) are the write side of budget.xlsx -- the structured API an
agent (the in-app AI Assistant's edit-budget mode, or an external coding
agent pointed at a user's data folder) uses to make a category-structure
change instead of ever hand-editing the workbook's cells directly. Every
write goes through these functions so a bad edit can't corrupt the
template's shape, and every one snapshots the previous file to
`<path>.bak` first.
"""

import shutil
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from pta_treasurer.builders import (
    NAVY_FILL, SUBHDR_FONT, TEAL_FILL, THIN_BORDER, WHITE,
)
from openpyxl.styles import Alignment, Font

TEMPLATE_HEADERS = [
    'Section', 'Item', 'QuickBooks Category Name(s)',
    'Last Year Actual', 'This Year Budget',
]

SHEET_NAMES = ('Income Budget', 'Expense Budget')


def _write_sheet(ws, title: str):
    ws.sheet_view.showGridLines = False
    for col, w in zip(['A', 'B', 'C', 'D', 'E'], [22, 28, 40, 16, 16]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:E1')
    c = ws['A1']
    c.value = title
    c.font = Font(name='Arial', bold=True, size=13, color=WHITE)
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    for ci, hdr in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=hdr)
        cell.font = SUBHDR_FONT
        cell.fill = TEAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[2].height = 30


def generate_template(path: Path) -> None:
    """
    Writes a blank budget workbook with 'Income Budget' / 'Expense Budget'
    sheets for the treasurer to fill in (or edit each fiscal year).
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = SHEET_NAMES[0]
    _write_sheet(ws1, 'INCOME BUDGET')

    ws2 = wb.create_sheet(SHEET_NAMES[1])
    _write_sheet(ws2, 'EXPENSE BUDGET')

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _read_budget_sheet(ws):
    """
    Returns (budget, qb_to_budget_map) for one sheet:
      budget: {section: {item: (last_year_actual, this_year_budget)}}
      qb_to_budget_map: {qb_category_name: item_name}
    """
    budget = {}
    qb_to_budget_map = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        section, item = row[0], row[1]
        if not section or not item:
            continue
        section = str(section).strip()
        item = str(item).strip()
        qb_names = str(row[2]).strip() if row[2] else ''
        last_year = float(row[3]) if row[3] not in (None, '') else 0.0
        this_year = float(row[4]) if row[4] not in (None, '') else 0.0

        budget.setdefault(section, {})[item] = (last_year, this_year)

        for qb_name in qb_names.split(','):
            qb_name = qb_name.strip()
            if qb_name:
                qb_to_budget_map[qb_name] = item

    return budget, qb_to_budget_map


def load_budget(path: Path):
    """
    Reads a budget workbook produced by generate_template (or hand-edited by
    the treasurer). Returns (income_budget, expense_budget, qb_to_budget_map)
    with the two budgets in the {section: {item: (last_yr, budget)}} shape
    that merge_actuals_into_budget expects, and qb_to_budget_map merged
    across both sheets.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    income_budget, income_map = _read_budget_sheet(wb[SHEET_NAMES[0]])
    expense_budget, expense_map = _read_budget_sheet(wb[SHEET_NAMES[1]])

    qb_to_budget_map = {**income_map, **expense_map}
    return income_budget, expense_budget, qb_to_budget_map


def map_actuals_to_budget_items(actuals_dict: dict, qb_to_budget_map: dict) -> dict:
    """
    Translates QuickBooks category names in actuals_dict to budget item
    names via qb_to_budget_map, summing entries that map to the same item.
    A QB category with no mapping keeps its own name unchanged (picked up
    later as an 'Other (from QuickBooks)' item by merge_actuals_into_budget,
    or simply ignored by callers -- like apply_dynamic_last_year -- that
    only care about items already in the budget).

    Args:
        actuals_dict:      {qb_category_name: [12 monthly actuals]}
        qb_to_budget_map: {qb_category_name: budget_item_name}

    Returns:
        {budget_item_name: [12 monthly actuals]}
    """
    mapped_actuals = {}
    for qb_name, vals in actuals_dict.items():
        budget_name = qb_to_budget_map.get(qb_name, qb_name)
        if budget_name in mapped_actuals:
            mapped_actuals[budget_name] = [
                mapped_actuals[budget_name][i] + vals[i] for i in range(12)
            ]
        else:
            mapped_actuals[budget_name] = list(vals)
    return mapped_actuals


def apply_dynamic_last_year(budget_dict: dict, prior_actuals_mapped: dict) -> dict:
    """
    Replaces each item's static 'last_year_actual' (originally hand-entered
    into budget.xlsx) with the real prior fiscal year's actual, computed
    from that year's own recorded history.

    Args:
        budget_dict:            {section: {item: (last_year_actual, this_year_budget)}}
        prior_actuals_mapped:   {budget_item_name: [12 monthly actuals]} --
                                 already translated via map_actuals_to_budget_items,
                                 for the PRIOR fiscal year specifically.

    Returns:
        budget_dict unchanged if prior_actuals_mapped is empty (no digital
        history exists at all for that prior fiscal year -- e.g. the very
        first fiscal year this app is used for -- so the static fallback
        value is kept). Otherwise, a new {section: {item: (dynamic_last_year,
        this_year_budget)}} -- an item with no matching prior-year actual
        becomes 0.0 (a genuinely zero prior year for that item), not the
        stale static value.
    """
    if not prior_actuals_mapped:
        return budget_dict

    return {
        section: {
            item: (round(sum(prior_actuals_mapped.get(item, [0.0] * 12)), 2), this_year)
            for item, (_, this_year) in items.items()
        }
        for section, items in budget_dict.items()
    }


def merge_actuals_into_budget(budget_dict: dict, actuals_dict: dict,
                               qb_to_budget_map: dict) -> dict:
    """
    Merges parsed QuickBooks actuals into a budget structure.

    Args:
        budget_dict:      {section: {item: (last_year_actual, this_year_budget)}}
        actuals_dict:      {qb_category_name: [12 monthly actuals]}
        qb_to_budget_map: {qb_category_name: budget_item_name}

    Returns:
        {section: {item: (last_year_actual, this_year_budget, [12 monthly actuals])}}
        plus an 'Other (from QuickBooks)' section for any QB category that
        doesn't map to a budget item.
    """
    mapped_actuals = map_actuals_to_budget_items(actuals_dict, qb_to_budget_map)

    result = {}
    all_budget_items = {item for section in budget_dict.values() for item in section}
    for section, items in budget_dict.items():
        result[section] = {}
        for item, (last_yr, budget) in items.items():
            result[section][item] = (last_yr, budget, mapped_actuals.get(item, [0.0] * 12))

    unmatched = [k for k in mapped_actuals if k not in all_budget_items]
    if unmatched:
        result['Other (from QuickBooks)'] = {
            item: (0.0, 0.0, mapped_actuals[item]) for item in unmatched
        }

    return result


def actuals_total_for_item(merged: dict, item: str) -> float:
    """Sums the 12 monthly actuals for `item` across a
    merge_actuals_into_budget() result. Returns 0.0 if the item isn't
    present at all (no actuals recorded for it this fiscal year) -- used
    by remove_item's safety check."""
    for items in merged.values():
        if item in items:
            _last_yr, _budget, actuals = items[item]
            return round(sum(actuals), 2)
    return 0.0


def _validate_sheet(sheet: str) -> None:
    if sheet not in SHEET_NAMES:
        raise ValueError(f'sheet must be one of {SHEET_NAMES}, got {sheet!r}')


def _sheet_rows(ws):
    """Yields (row_idx, section, item) for every non-blank data row."""
    for row_idx in range(3, ws.max_row + 1):
        section = ws.cell(row=row_idx, column=1).value
        item = ws.cell(row=row_idx, column=2).value
        if section and item:
            yield row_idx, str(section).strip(), str(item).strip()


def _find_item_row(ws, item: str) -> int | None:
    for row_idx, _section, it in _sheet_rows(ws):
        if it == item:
            return row_idx
    return None


def _first_empty_row(ws) -> int:
    last = 2
    for row_idx, _section, _item in _sheet_rows(ws):
        last = row_idx
    return last + 1


def _backup(path: Path) -> None:
    """Snapshots the pre-edit file to <stem>.bak<suffix> (e.g.
    budget.bak.xlsx) -- keeping the real extension, unlike a bare .bak
    suffix, means the backup stays directly openable in Excel or via
    load_budget() if a treasurer needs to recover a prior version."""
    path = Path(path)
    if path.exists():
        shutil.copy2(path, path.with_name(f'{path.stem}.bak{path.suffix}'))


def add_item(path: Path, sheet: str, section: str, item: str, qb_names=(),
             last_year_actual: float = 0.0, budget: float = 0.0) -> None:
    """Appends a new budget line item. Raises ValueError if `item` already
    exists anywhere in `sheet` -- use rename_item/set_budget_amount to
    change an existing one instead of adding a duplicate."""
    _validate_sheet(sheet)
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    if _find_item_row(ws, item) is not None:
        raise ValueError(
            f"'{item}' already exists in {sheet} -- use a different name, "
            f"or edit the existing item instead of adding a duplicate.")
    row_idx = _first_empty_row(ws)
    ws.cell(row=row_idx, column=1, value=section)
    ws.cell(row=row_idx, column=2, value=item)
    ws.cell(row=row_idx, column=3, value=', '.join(qb_names))
    ws.cell(row=row_idx, column=4, value=round(last_year_actual, 2))
    ws.cell(row=row_idx, column=5, value=round(budget, 2))
    _backup(path)
    wb.save(path)


def remove_item(path: Path, sheet: str, item: str,
                 current_actuals_total: float | None = None,
                 force: bool = False) -> None:
    """Deletes a budget line item's row. Refuses (ValueError) when
    current_actuals_total is given and nonzero, unless force=True -- this
    stops an edit from silently dropping real transaction history from
    the budget view. Pass current_actuals_total=None (the default) to
    skip the check entirely, e.g. when no report has been generated yet
    to compute it from."""
    if current_actuals_total not in (None, 0.0) and not force:
        raise ValueError(
            f"'{item}' has ${current_actuals_total:,.2f} in actual "
            f"transactions recorded this fiscal year -- removing it would "
            f"drop that history from the budget view. Pass force=True if "
            f"you're sure.")
    _validate_sheet(sheet)
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    row_idx = _find_item_row(ws, item)
    if row_idx is None:
        raise ValueError(f"'{item}' not found in {sheet}.")
    ws.delete_rows(row_idx)
    _backup(path)
    wb.save(path)


def move_item(path: Path, sheet: str, item: str, to_section: str) -> None:
    """Re-keys an item to a different section within the same sheet -- a
    pure display-grouping change; the item's actuals/history are
    untouched since those are looked up by item name, not section."""
    _validate_sheet(sheet)
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    row_idx = _find_item_row(ws, item)
    if row_idx is None:
        raise ValueError(f"'{item}' not found in {sheet}.")
    ws.cell(row=row_idx, column=1, value=to_section)
    _backup(path)
    wb.save(path)


def rename_item(path: Path, sheet: str, old_name: str, new_name: str) -> None:
    """Renames an item in place. Its QuickBooks Category Name(s) mapping
    lives on the same row, so it moves with the rename automatically --
    no separate map to update."""
    _validate_sheet(sheet)
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    if _find_item_row(ws, new_name) is not None:
        raise ValueError(f"'{new_name}' already exists in {sheet}.")
    row_idx = _find_item_row(ws, old_name)
    if row_idx is None:
        raise ValueError(f"'{old_name}' not found in {sheet}.")
    ws.cell(row=row_idx, column=2, value=new_name)
    _backup(path)
    wb.save(path)


def map_qb_category(path: Path, sheet: str, item: str, qb_name: str) -> None:
    """Adds qb_name to item's comma-separated QuickBooks Category Name(s)
    list (deduped, order-preserving) -- the 'QuickBooks renamed this
    category' case."""
    _validate_sheet(sheet)
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    row_idx = _find_item_row(ws, item)
    if row_idx is None:
        raise ValueError(f"'{item}' not found in {sheet}.")
    cell = ws.cell(row=row_idx, column=3)
    existing = [n.strip() for n in str(cell.value or '').split(',') if n.strip()]
    if qb_name not in existing:
        existing.append(qb_name)
    cell.value = ', '.join(existing)
    _backup(path)
    wb.save(path)


def set_budget_amount(path: Path, sheet: str, item: str, amount: float) -> None:
    """Updates an item's This Year Budget figure."""
    _validate_sheet(sheet)
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet]
    row_idx = _find_item_row(ws, item)
    if row_idx is None:
        raise ValueError(f"'{item}' not found in {sheet}.")
    ws.cell(row=row_idx, column=5, value=round(amount, 2))
    _backup(path)
    wb.save(path)


def describe_edit(action: dict) -> str:
    """Pure. One-line, human-readable preview of a parsed edit action
    (see ai_assistant.parse_edit_action), for a preview-diff-then-confirm
    UI. Raises ValueError for an unrecognized action name -- should never
    happen in practice since parse_edit_action validates the action name
    first and downgrades anything else to 'clarify'."""
    a = action.get('action')
    if a == 'add_item':
        qb = (f" (QuickBooks: {', '.join(action['qb_names'])})"
              if action.get('qb_names') else '')
        return (f"Add new item '{action['item']}' to {action['sheet']} -> "
                f"{action['section']} section, budget "
                f"${action.get('budget', 0.0):,.2f}{qb}")
    if a == 'remove_item':
        return f"Remove '{action['item']}' from {action['sheet']}"
    if a == 'move_item':
        return (f"Move '{action['item']}' to the '{action['to_section']}' "
                f"section in {action['sheet']}")
    if a == 'rename_item':
        return (f"Rename '{action['old_name']}' to '{action['new_name']}' "
                f"in {action['sheet']}")
    if a == 'map_qb_category':
        return (f"Map QuickBooks category '{action['qb_name']}' to "
                f"'{action['item']}' in {action['sheet']}")
    if a == 'set_budget_amount':
        return (f"Set '{action['item']}' budget to "
                f"${action['amount']:,.2f} in {action['sheet']}")
    raise ValueError(f'Unknown edit action: {a!r}')


def apply_edit(path: Path, action: dict,
                current_actuals_total: float | None = None) -> None:
    """Dispatches a parsed edit action to the matching write function
    above -- the single call site both the GUI's Apply button and an
    external agent use, so every write goes through the same validation
    regardless of caller. current_actuals_total is only consulted for
    remove_item (see remove_item's docstring)."""
    a = action.get('action')
    if a == 'add_item':
        add_item(path, action['sheet'], action['section'], action['item'],
                  qb_names=action.get('qb_names') or (),
                  last_year_actual=action.get('last_year_actual', 0.0),
                  budget=action.get('budget', 0.0))
    elif a == 'remove_item':
        remove_item(path, action['sheet'], action['item'],
                     current_actuals_total=current_actuals_total)
    elif a == 'move_item':
        move_item(path, action['sheet'], action['item'], action['to_section'])
    elif a == 'rename_item':
        rename_item(path, action['sheet'], action['old_name'], action['new_name'])
    elif a == 'map_qb_category':
        map_qb_category(path, action['sheet'], action['item'], action['qb_name'])
    elif a == 'set_budget_amount':
        set_budget_amount(path, action['sheet'], action['item'], action['amount'])
    else:
        raise ValueError(f'Unknown edit action: {a!r}')
