"""
pipeline.py
Orchestrates a single month's report generation: locate input files, parse
them, roll them into history, merge history into the budget, and build the
output workbook. This is the one code path both the CLI and the GUI call.
"""

import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

import openpyxl

from pta_treasurer.budget_io import (
    apply_dynamic_last_year, load_budget, map_actuals_to_budget_items,
    merge_actuals_into_budget,
)
from pta_treasurer.builders import (
    FISCAL_MONTHS, build_budget, build_credits_sheet, build_debits_sheet,
    build_givebacks, build_manifest, build_memberhub_summary_sheet,
    build_treasurer, build_ytd_summary, build_ytd_summary_compact,
)
from pta_treasurer.config import (
    MONTH_NAMES, OrgConfig, detect_month_from_filename, detect_month_from_pdf,
    fiscal_year_start_calendar_year, month_name_to_fiscal_index,
)
from pta_treasurer.parsers import (
    consolidate_givebacks_payouts, extract_payout_id, find_bank_statement_month,
    match_credits_to_bank_statement, parse_chase_pdf, parse_givebacks_files,
    parse_quickbooks_detail,
)


# One-time correction: FY2026-27's July 2026 QuickBooks expenses include
# late bills for FY2025-26 activities. Rather than a frozen snapshot of
# dollar amounts, this pulls July 2026's REAL recorded actuals (whatever
# they turn out to be, including later bookkeeping corrections) for every
# expense item except the excluded ones below, and folds them into
# FY2025-26's ("last year's") actual total in addition to their normal
# appearance in FY2026-27's own JULY column. Only applies when computing
# FY2025-26 as the PRIOR fiscal year (i.e. while generating any FY2026-27
# report) -- inert for every other fiscal year, and safe to delete once
# FY2025-26 stops being anyone's "last year".
_JULY_2026_SPILLOVER_PRIOR_FY_START = 2025
_JULY_2026_SPILLOVER_EXCLUDED_ITEMS = {'Accounting Quickbooks'}

# One-time correction: QuickBooks recorded a single $901.00 raw '5th grade
# T-shirts' category for October 2025, but it actually covered two separate
# real orders -- a $125.00 staff order and a $776.00 grade order. A raw QB
# category can only route to one budget line via qb_to_budget_map, so this
# splits the raw actual itself (fiscal_index 3 = October) before anything
# downstream maps or sums it. Only applies to FY2025-26's own raw actuals
# (as the current fiscal year, or as the prior fiscal year for a FY2026-27
# report's "Last Year Actual"); inert for every other fiscal year, and safe
# to delete once no longer relevant.
_OCT_2025_TSHIRT_STAFF_SPLIT = 125.00


def _apply_oct_2025_tshirt_split(expense_actuals: dict, fiscal_year_start: int) -> dict:
    if fiscal_year_start != 2025:
        return expense_actuals
    raw = expense_actuals.get('5th grade T-shirts')
    if not raw or not raw[3]:
        return expense_actuals
    expense_actuals = dict(expense_actuals)

    grade_monthly = list(raw)
    grade_monthly[3] = round(grade_monthly[3] - _OCT_2025_TSHIRT_STAFF_SPLIT, 2)
    expense_actuals['5th grade T-shirts'] = grade_monthly

    staff_monthly = list(expense_actuals.get('5th Staff T-Shirts', [0.0] * 12))
    staff_monthly[3] = round(staff_monthly[3] + _OCT_2025_TSHIRT_STAFF_SPLIT, 2)
    expense_actuals['5th Staff T-Shirts'] = staff_monthly
    return expense_actuals


@dataclass
class RunResult:
    output_path: Path
    warnings: list = field(default_factory=list)
    qb: dict = field(default_factory=dict)
    bank: dict = field(default_factory=dict)
    givebacks: list = field(default_factory=list)
    income_merged: dict = field(default_factory=dict)
    expense_merged: dict = field(default_factory=dict)


def _history_base_dir(data_dir: Path) -> Path:
    return Path(data_dir) / 'data' / 'history'


def _history_dir(data_dir: Path, fiscal_year_start: int) -> Path:
    """The one fiscal year's history subfolder, e.g. data/history/2025_to_2026/
    -- history is partitioned per fiscal year so that a calendar month's
    fiscal_index (0-11, which repeats every fiscal year) never collides
    across different fiscal years' data on disk."""
    return _history_base_dir(data_dir) / f'{fiscal_year_start}_to_{fiscal_year_start + 1}'


def _locate_month_files(month_folder: Path, month: str):
    """
    Finds the QuickBooks CSV, bank statement PDF, and Givebacks CSVs for a
    month folder. Raises FileNotFoundError if a required file is missing.
    """
    all_csv = sorted(month_folder.glob('*.csv'))
    qb_files = [f for f in all_csv
                if month.lower() in f.name.lower() and 'givebacks' not in f.name.lower()]
    if not qb_files:
        raise FileNotFoundError(
            f'No QuickBooks file found for {month} in {month_folder}\n'
            f'Files available: {[f.name for f in all_csv]}'
        )
    qb_file = qb_files[-1]

    pdf_files = sorted(month_folder.glob('*.pdf'))
    if not pdf_files:
        raise FileNotFoundError(f'No bank statement PDF found in {month_folder}')
    matching_pdf = [f for f in pdf_files if month.lower() in f.name.lower()]
    bank_file = matching_pdf[-1] if matching_pdf else pdf_files[-1]

    gb_folder = month_folder / 'givebacks'
    gb_file_info = []
    if gb_folder.exists():
        for f in sorted(gb_folder.glob('*.csv')):
            lbl, idx = detect_month_from_filename(f)
            gb_file_info.append((f, lbl or month.capitalize(), idx))

    return qb_file, bank_file, gb_file_info


def _save_history_entry(data_dir: Path, month_label: str, fiscal_idx: int,
                         qb: dict, givebacks: list) -> None:
    fy_start = fiscal_year_start_calendar_year(month_label)
    history_dir = _history_dir(data_dir, fy_start)
    history_dir.mkdir(parents=True, exist_ok=True)
    entry = {
        'month_label':                month_label,
        'fiscal_index':                fiscal_idx,
        'income':                      qb['income'],
        'expenses':                    qb['expenses'],
        'income_total':                qb['income_total'],
        'expense_total':               qb['expense_total'],
        'net_income':                  qb['net_income'],
        'pass_through_income_total':   qb.get('pass_through_income_total', 0.0),
        'pass_through_expense_total':  qb.get('pass_through_expense_total', 0.0),
        'pass_through_net':            qb.get('pass_through_net', 0.0),
        'givebacks_total':             sum(g['total'] for g in givebacks),
        'generated_at':                datetime.now().isoformat(),
    }
    safe_month = month_label.replace(' ', '_')
    (history_dir / f'{safe_month}.json').write_text(json.dumps(entry, indent=2))


def compute_pass_through_balance_held(data_dir: Path, balance_forward: float,
                                       before_fiscal_year_start: int | None = None) -> float:
    """Sums pass_through_net across every fiscal year's history entry (i.e.
    data_dir/data/history/*/*.json, across ALL fiscal-year subfolders), in
    true calendar order (not filename order), starting from balance_forward
    -- so the running balance survives fiscal-year boundaries. This is
    intentionally left unscoped to one fiscal year: pass-through fund money
    (e.g. a fundraiser's proceeds) carries over indefinitely and isn't the
    org's own money, so its running balance must span fiscal years even
    though every other actuals lookup in this module is scoped to one.
    Entries predating this feature simply contribute 0.0.

    before_fiscal_year_start: if given, only sums entries strictly BEFORE
    that fiscal year's July 1 (i.e. everything held as of the start of
    that fiscal year, before its own transactions) -- used by
    get_fiscal_year_balance_forward() to exclude pass-through money from
    Balance Forward. None (the default) sums everything ever recorded, for
    "balance held right now" callers."""
    history_dir = _history_base_dir(data_dir)
    if not history_dir.exists():
        return balance_forward

    dated = []
    for hf in history_dir.glob('*/*.json'):
        try:
            entry = json.loads(hf.read_text())
            dt = datetime.strptime(entry['month_label'], '%B %Y')
            dated.append((dt, entry))
        except (json.JSONDecodeError, OSError, KeyError, ValueError):
            continue
    if before_fiscal_year_start is not None:
        cutoff = datetime(before_fiscal_year_start, 7, 1)
        dated = [(dt, entry) for dt, entry in dated if dt < cutoff]
    dated.sort(key=lambda t: t[0])

    running = balance_forward
    for _, entry in dated:
        running += entry.get('pass_through_net', 0.0)
    return running


def load_all_actuals(data_dir: Path, fiscal_year_start: int):
    """
    Loads one fiscal year's history entries (from its own
    data/history/{fiscal_year_start}_to_{fiscal_year_start+1}/ subfolder)
    into 12-slot-per-item actuals arrays, indexed by each entry's
    fiscal_index. Returns (income_actuals, expense_actuals).

    History is partitioned on disk per fiscal year specifically so that a
    calendar month's fiscal_index (0-11, which repeats every fiscal year)
    can never collide with another fiscal year's data -- both July_2025 and
    July_2026 share fiscal_index 0, and reading from a shared/flat directory
    previously let the alphabetically-later file silently overwrite the
    correct one (confirmed happening for real). fiscal_year_start is
    required -- there is no "everything ever recorded" mode; call this once
    per fiscal year you need (e.g. once for the current year's own actuals,
    once for the prior year's, when computing dynamic "last year" values).
    """
    history_dir = _history_dir(data_dir, fiscal_year_start)
    if not history_dir.exists():
        return {}, {}

    income_actuals = defaultdict(lambda: [0.0] * 12)
    expense_actuals = defaultdict(lambda: [0.0] * 12)

    for hf in sorted(history_dir.glob('*.json')):
        try:
            entry = json.loads(hf.read_text())
        except (json.JSONDecodeError, OSError):
            continue
        idx = entry.get('fiscal_index')
        if idx is None:
            continue
        for item, val in entry.get('income', {}).items():
            income_actuals[item][idx] = val
        for item, val in entry.get('expenses', {}).items():
            expense_actuals[item][idx] = val

    return dict(income_actuals), dict(expense_actuals)


def get_fiscal_year_balance_forward(config: OrgConfig, data_dir: Path,
                                     fiscal_year_start: int, month_label: str,
                                     bank: dict) -> tuple:
    """
    Bank balance as of the start of the given fiscal year (July 1) -- the
    YTD Summary's 'Balance Forward' figure. Reads it from that fiscal
    year's own July Chase statement so it's always correct no matter which
    month is currently being processed, instead of a single static
    config.balance_forward value that had to be hand-edited every fiscal
    year and silently went stale for every month in between (confirmed
    showing $0.00 for a FY2026-27 report in production).

    If the month being processed IS that fiscal year's July, the
    already-parsed `bank` object is reused directly rather than re-reading
    the PDF. Falls back to config.balance_forward (the original
    manual-entry value, used as-is with no further adjustment) only when
    no digital July statement exists for that fiscal year at all -- e.g.
    the very first fiscal year this app is used for, before any input
    folder for it exists.

    When a pass-through fund is configured, the raw bank balance is
    reduced by whatever pass-through money (e.g. a fundraiser's proceeds)
    was already sitting in the account as of that July 1 -- Balance
    Forward should represent the org's OWN money carried forward, the same
    "PTA Money" definition already used for Current Balance, not money
    that passes through the account but isn't the org's.

    Returns (balance_forward, warning_or_None).
    """
    if month_label == f'July {fiscal_year_start}':
        raw_balance = bank['beginning_balance']
    else:
        july_folder = Path(data_dir) / 'input' / f'July_{fiscal_year_start}'
        pdf_files = sorted(july_folder.glob('*.pdf')) if july_folder.exists() else []
        if not pdf_files:
            return config.balance_forward, None
        try:
            raw_balance = parse_chase_pdf(pdf_files[-1])['beginning_balance']
        except Exception as e:
            return config.balance_forward, (
                f'Could not parse July {fiscal_year_start} Chase PDF for Balance Forward '
                f'({e}) -- falling back to the manually-entered value.'
            )

    if config.pass_through_fund_name:
        passthrough_held = compute_pass_through_balance_held(
            data_dir, config.pass_through_fund_balance_forward,
            before_fiscal_year_start=fiscal_year_start)
        raw_balance -= passthrough_held

    return raw_balance, None


def run_month(config: OrgConfig, data_dir: Path, month: str, year: str) -> RunResult:
    """
    Generates the monthly treasurer report workbook for `month year`.

    Reads data_dir/input/{Month}_{Year}/ for the QuickBooks CSV, Chase PDF,
    and (optional) givebacks/*.csv exports; reads data_dir/budget.xlsx for
    the budget config; writes
    data_dir/data/history/{fy_start}_to_{fy_end}/{Month}_{Year}.json;
    writes
    data_dir/output/Treasurer_Report_{fy_start}_to_{fy_end}_{Month}_{Year}.xlsx.
    """
    data_dir = Path(data_dir)
    warnings = []

    fiscal_idx = month_name_to_fiscal_index(month)
    if fiscal_idx is None:
        raise ValueError(f'Unrecognized month name: {month!r}')
    month_label = f'{month.capitalize()} {year}'

    month_folder = data_dir / 'input' / f'{month.capitalize()}_{year}'
    if not month_folder.exists():
        raise FileNotFoundError(f'No input folder found at {month_folder}')

    qb_file, bank_file, gb_file_info = _locate_month_files(month_folder, month)

    # Cross-validate detected months against the requested month — catches
    # the wrong statement being dropped into the wrong month's folder.
    qb_detected, _ = detect_month_from_filename(qb_file)
    bank_detected, _ = detect_month_from_pdf(bank_file)
    for label, detected in [('QuickBooks file', qb_detected), ('Bank statement', bank_detected)]:
        if detected and not detected.lower().startswith(month.lower()):
            raise ValueError(
                f'{label} ({detected}) does not match requested month ({month_label})'
            )

    pass_through_categories = None
    if config.pass_through_fund_name:
        pass_through_categories = {
            c.strip().upper() for c in config.pass_through_fund_categories.split(',') if c.strip()
        }
    qb = parse_quickbooks_detail(month_folder, month, year,
                                  pass_through_categories=pass_through_categories,
                                  pass_through_fund_name=config.pass_through_fund_name)
    bank = parse_chase_pdf(bank_file)

    if gb_file_info:
        givebacks = parse_givebacks_files(gb_file_info)
    else:
        givebacks = []
        warnings.append(
            f'No Givebacks files found for {month_label} — '
            f'Giveback Reconciliation sheet will be empty.'
        )

    _save_history_entry(data_dir, month_label, fiscal_idx, qb, givebacks)

    budget_path = data_dir / 'budget.xlsx'
    if not budget_path.exists():
        raise FileNotFoundError(
            f'No budget template found at {budget_path} — '
            f'generate one with budget_io.generate_template() first.'
        )
    income_budget, expense_budget, qb_to_budget_map = load_budget(budget_path)

    target_fy_start = fiscal_year_start_calendar_year(month_label)
    prior_fy_start = target_fy_start - 1

    income_actuals, expense_actuals = load_all_actuals(
        data_dir, fiscal_year_start=target_fy_start)
    expense_actuals = _apply_oct_2025_tshirt_split(expense_actuals, target_fy_start)
    target_expense_mapped = map_actuals_to_budget_items(expense_actuals, qb_to_budget_map)

    prior_income_actuals, prior_expense_actuals = load_all_actuals(
        data_dir, fiscal_year_start=prior_fy_start)
    prior_expense_actuals = _apply_oct_2025_tshirt_split(prior_expense_actuals, prior_fy_start)
    prior_income_mapped = map_actuals_to_budget_items(prior_income_actuals, qb_to_budget_map)
    prior_expense_mapped = map_actuals_to_budget_items(prior_expense_actuals, qb_to_budget_map)

    if prior_fy_start == _JULY_2026_SPILLOVER_PRIOR_FY_START:
        for item, vals in target_expense_mapped.items():
            if item in _JULY_2026_SPILLOVER_EXCLUDED_ITEMS:
                continue
            july_amount = vals[0]  # fiscal_index 0 == July
            if not july_amount:
                continue
            existing = list(prior_expense_mapped.get(item, [0.0] * 12))
            existing[0] += july_amount
            prior_expense_mapped[item] = existing

    income_budget = apply_dynamic_last_year(income_budget, prior_income_mapped)
    expense_budget = apply_dynamic_last_year(expense_budget, prior_expense_mapped)

    income_merged = merge_actuals_into_budget(income_budget, income_actuals, qb_to_budget_map)
    expense_merged = merge_actuals_into_budget(expense_budget, expense_actuals, qb_to_budget_map)

    wb = openpyxl.Workbook()

    pass_through = None
    if config.pass_through_fund_name:
        balance_held = compute_pass_through_balance_held(
            data_dir, config.pass_through_fund_balance_forward)
        pass_through = {
            'fund_name':      config.pass_through_fund_name,
            'income_total':   qb.get('pass_through_income_total', 0.0),
            'expense_total':  qb.get('pass_through_expense_total', 0.0),
            'net':            qb.get('pass_through_net', 0.0),
            'balance_held':   balance_held,
        }

    ws1 = wb.active
    ws1.title = 'Treasurer Report'
    build_treasurer(ws1, qb, bank, month_label, config.org_name, pass_through=pass_through)

    balance_forward, balance_forward_warning = get_fiscal_year_balance_forward(
        config, data_dir, target_fy_start, month_label, bank)
    if balance_forward_warning:
        warnings.append(balance_forward_warning)

    # YTD Summary comes right after the Treasurer Report -- the overview a
    # treasurer wants first -- rather than last among the tabs. The compact
    # (5-column) layout is the default; build_ytd_summary (the original
    # 9-column version) stays available for on-demand/programmatic use.
    ws2 = wb.create_sheet('YTD Summary')
    build_ytd_summary_compact(ws2, income_merged, expense_merged, config.org_name,
                               month_label, fiscal_idx, FISCAL_MONTHS, target_fy_start,
                               bank=bank, balance_forward=balance_forward,
                               pass_through=pass_through)

    ws3 = wb.create_sheet('Income Budget vs Actuals')
    build_budget(ws3, 'Budget vs Actuals - Income', income_merged,
                 config.org_name, FISCAL_MONTHS, fiscal_idx, target_fy_start, show_pl=False)

    ws4 = wb.create_sheet('Expense Budget vs Actuals')
    build_budget(ws4, 'Budget vs Actuals - Expenses', expense_merged,
                 config.org_name, FISCAL_MONTHS, fiscal_idx, target_fy_start,
                 show_pl=True, income_merged=income_merged)

    ws5 = wb.create_sheet('Giveback Reconciliation')
    build_givebacks(ws5, givebacks, bank, config.org_name)

    ws6 = wb.create_sheet('File Manifest')
    build_manifest(ws6, month_folder, config.org_name, month_label, fiscal_idx, FISCAL_MONTHS)

    output_dir = data_dir / 'output'
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_month = month_label.replace(' ', '_')
    output_path = (
        output_dir
        / f'Treasurer_Report_{target_fy_start}_to_{target_fy_start + 1}_{safe_month}.xlsx'
    )
    wb.save(output_path)

    return RunResult(output_path=output_path, warnings=warnings,
                      qb=qb, bank=bank, givebacks=givebacks,
                      income_merged=income_merged, expense_merged=expense_merged)


@dataclass
class LedgerResult:
    output_path: Path
    warnings: list = field(default_factory=list)


def _find_month_folders(data_dir: Path):
    """Every data_dir/input/{Month}_{Year} folder with a QuickBooks CSV, in
    true calendar order (so a fiscal year spanning two calendar years - e.g.
    July 2025 through June 2026 - comes out in the right sequence). Returns
    [(month_name, year, cal_idx, folder), ...]."""
    base = Path(data_dir) / 'input'
    found = []
    if not base.exists():
        return found
    for folder in sorted(base.glob('*_*')):
        if not folder.is_dir():
            continue
        m = re.match(r'^([A-Za-z]+)_(\d{4})$', folder.name)
        if not m:
            continue
        month_name, year = m.group(1), m.group(2)
        cal_idx = MONTH_NAMES.get(month_name.lower())
        if cal_idx is None:
            continue
        qb_files = [f for f in folder.glob('*.csv') if 'givebacks' not in f.name.lower()]
        if not qb_files:
            continue
        found.append((month_name, year, cal_idx, folder))
    found.sort(key=lambda t: (int(t[1]), t[2]))
    return found


def list_available_fiscal_years(data_dir: Path) -> list:
    """Every fiscal-year start calendar year with at least one month folder
    under data_dir/input/, most recent first -- for a fiscal-year picker.
    E.g. [2026, 2025] if both FY2025-26 and FY2026-27 have data."""
    month_folders = _find_month_folders(data_dir)
    years = {
        fiscal_year_start_calendar_year(f'{mn.capitalize()} {y}')
        for mn, y, _, _ in month_folders
    }
    return sorted(years, reverse=True)


def build_debits_credits_ledger(config: OrgConfig, data_dir: Path,
                                 fiscal_year_start: int | None = None) -> LedgerResult:
    """
    Rebuilds a whole-fiscal-year "Debits & Credits" check-register-style
    ledger (Credits/Debits/MemberHub_Summary sheets) from every
    data_dir/input/{Month}_{Year} folder in the given fiscal year.
    fiscal_year_start=None defaults to the fiscal year containing today's
    date -- pass an explicit year (e.g. from list_available_fiscal_years())
    to build a past fiscal year instead, since "today" is a poor default
    once a new fiscal year has already started and you want to review the
    one that just closed. Independent of run_month()/a single month's
    report -- always rebuilds fully from scratch (safe/idempotent), and
    resilient to a single bad month (warns and continues rather than
    aborting the whole ledger).
    """
    data_dir = Path(data_dir)
    warnings = []

    month_folders = _find_month_folders(data_dir)

    if fiscal_year_start is None:
        today = date.today()
        today_label = f'{today.strftime("%B")} {today.year}'
        fiscal_year_start = fiscal_year_start_calendar_year(today_label)
    current_fy_start = fiscal_year_start
    month_folders = [
        (mn, y, idx, folder) for mn, y, idx, folder in month_folders
        if fiscal_year_start_calendar_year(f'{mn.capitalize()} {y}') == current_fy_start
    ]

    if not month_folders:
        raise FileNotFoundError(
            f'No month folders with QuickBooks data found for fiscal year '
            f'{current_fy_start}-{current_fy_start + 1} under {data_dir / "input"}'
        )

    fiscal_order = [f'{mn.capitalize()} {y}' for mn, y, _, _ in month_folders]

    # Parse every month's Chase PDF up front -- the "next month" lag lookup
    # in match_credits_to_bank_statement/find_bank_statement_month needs
    # every month's bank data available, not just a forward single pass.
    bank_by_month = {}
    for month_name, year, _, folder in month_folders:
        month_label = f'{month_name.capitalize()} {year}'
        pdf_files = sorted(folder.glob('*.pdf'))
        if not pdf_files:
            warnings.append(
                f'No Chase PDF for {month_label} — bank statement matching skipped for this month')
            continue
        try:
            bank_by_month[month_label] = parse_chase_pdf(pdf_files[-1])
        except Exception as e:
            warnings.append(f'Could not parse Chase PDF for {month_label}: {e}')

    credits_by_month = []
    debits_by_month = []
    givebacks_by_month = []
    payout_id_locations = {}  # payout id -> [(month_label, filename), ...]

    for month_name, year, _, folder in month_folders:
        month_label = f'{month_name.capitalize()} {year}'
        try:
            qb = parse_quickbooks_detail(folder, month_name.capitalize(), year)
        except Exception as e:
            warnings.append(f'Could not parse QuickBooks for {month_label}: {e}')
            continue

        credit_txns = [t for t in qb['transactions'] if t['is_income']]
        debit_txns = [t for t in qb['transactions'] if not t['is_income']]

        match_credits_to_bank_statement(credit_txns, month_label, bank_by_month, fiscal_order)
        for t in debit_txns:
            t['bank_statement_month'] = find_bank_statement_month(
                t, month_label, bank_by_month, fiscal_order)

        gb_folder = folder / 'givebacks'
        gb_files = sorted(gb_folder.rglob('*.csv')) if gb_folder.exists() else []
        payouts = []
        for f in gb_files:
            payout_id = extract_payout_id(f.name)
            if payout_id:
                payout_id_locations.setdefault(payout_id, []).append((month_label, f.name))
            try:
                items = parse_givebacks_files([(f, month_label, 0)])
                payouts.append({
                    'total': sum(it['total'] for it in items),
                    'items': items,
                    'source_file': f.name,
                })
            except Exception as e:
                warnings.append(f'Could not parse Givebacks payout {f.name}: {e}')

        consolidated_credits, payout_dates = consolidate_givebacks_payouts(credit_txns, payouts)
        credits_by_month.append((month_label, consolidated_credits))
        debits_by_month.append((month_label, debit_txns))

        if payouts:
            month_payouts = [
                {'date': payout_dates.get(i), 'total': p['total'], 'items': p['items']}
                for i, p in enumerate(payouts)
            ]
            givebacks_by_month.append((month_label, month_payouts))

    duplicate_payouts = {pid: locs for pid, locs in payout_id_locations.items() if len(locs) > 1}
    for pid, locs in duplicate_payouts.items():
        where = ', '.join(f'{month} ({fname})' for month, fname in locs)
        warnings.append(
            f'Givebacks payout {pid} found in more than one month: {where} — '
            f'remove the duplicate copy from whichever month it does NOT actually belong to.')

    qb_to_budget_map = {}
    budget_path = data_dir / 'budget.xlsx'
    if budget_path.exists():
        try:
            _, _, qb_to_budget_map = load_budget(budget_path)
        except Exception as e:
            warnings.append(f'Could not load budget.xlsx for BUDGET LINE mapping: {e}')

    fy_start = current_fy_start
    fy_end = current_fy_start + 1
    output_dir = data_dir / 'output'
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f'Debits_and_Credits_{fy_start}_to_{fy_end}.xlsx'

    wb = openpyxl.Workbook()
    ws_credits = wb.active
    ws_credits.title = 'Credits'
    build_credits_sheet(ws_credits, credits_by_month, config.org_name, qb_to_budget_map)

    ws_debits = wb.create_sheet('Debits')
    build_debits_sheet(ws_debits, debits_by_month, config.org_name, qb_to_budget_map)

    ws_memberhub = wb.create_sheet('MemberHub_Summary')
    build_memberhub_summary_sheet(ws_memberhub, givebacks_by_month, config.org_name)

    wb.save(output_path)

    return LedgerResult(output_path=output_path, warnings=warnings)
