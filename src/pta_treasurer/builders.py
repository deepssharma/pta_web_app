"""
builders.py
All Excel sheet builder functions for the PTA Treasurer Report Generator.
Requires openpyxl.
"""

from datetime import datetime
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colours ───────────────────────────────────────────────────────────────────
NAVY   = '1F3864'; TEAL   = '2E75B6'; LTBLUE = 'BDD7EE'
GOLD   = 'FFD966'; WHITE  = 'FFFFFF'; LGREY  = 'F2F2F2'
GREEN  = 'E2EFDA'; RED_BG = 'FCE4D6'

SUBHDR_FONT = Font(name='Arial', bold=True, color=WHITE, size=10)
BODY_FONT   = Font(name='Arial', size=10)
BOLD_FONT   = Font(name='Arial', bold=True, size=10)
TOTAL_FONT  = Font(name='Arial', bold=True, size=10, color=NAVY)

NAVY_FILL   = PatternFill('solid', fgColor=NAVY)
TEAL_FILL   = PatternFill('solid', fgColor=TEAL)
LTBLUE_FILL = PatternFill('solid', fgColor=LTBLUE)
GOLD_FILL   = PatternFill('solid', fgColor=GOLD)
LGREY_FILL  = PatternFill('solid', fgColor=LGREY)
GREEN_FILL  = PatternFill('solid', fgColor=GREEN)
RED_FILL    = PatternFill('solid', fgColor=RED_BG)

THIN        = Side(style='thin',   color='AAAAAA')
MED         = Side(style='medium', color=NAVY)
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

# Brackets a variable-length group of rows (e.g. a Credits-sheet "Bank
# Deposit" band plus its nested category breakdown) with a bold outer
# border and thin seams between rows inside it, so the group reads as
# one unit rather than a run of ordinary transaction rows.
GROUP_TOP_BORDER  = Border(left=MED, right=MED, top=MED,  bottom=THIN)
GROUP_MID_BORDER  = Border(left=MED, right=MED, top=THIN, bottom=THIN)
GROUP_LAST_BORDER = Border(left=MED, right=MED, top=THIN, bottom=MED)
MONEY_FMT   = '$#,##0.00_);($#,##0.00)'

FISCAL_MONTHS = ['JULY','AUG','SEPT','OCT','NOV','DEC',
                 'JAN','FEB','MAR','APR','MAY','JUNE']


# ── Shared helpers ────────────────────────────────────────────────────────────

def _sec_hdr(ws, label, row, n=4):
    ws.merge_cells(f'A{row}:{get_column_letter(n)}{row}')
    c = ws[f'A{row}']; c.value = label
    c.font = Font(name='Arial', bold=True, size=11, color=WHITE)
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 20
    return row + 1


def _col_hdrs(ws, row, labels):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=i+1, value=lbl)
        c.font = SUBHDR_FONT; c.fill = TEAL_FILL; c.border = THIN_BORDER
        c.alignment = Alignment(
            horizontal='center' if i > 0 else 'left',
            vertical='center', indent=1 if i == 0 else 0)
    ws.row_dimensions[row].height = 18
    return row + 1


def _data_row(ws, row, label, amount, shade=False):
    fill = LGREY_FILL if shade else PatternFill()
    ws[f'A{row}'].value = label; ws[f'A{row}'].font = BODY_FONT
    ws[f'A{row}'].fill = fill;   ws[f'A{row}'].border = THIN_BORDER
    ws[f'A{row}'].alignment = Alignment(indent=2)
    ws[f'B{row}'].value = amount; ws[f'B{row}'].font = BODY_FONT
    ws[f'B{row}'].fill = fill;    ws[f'B{row}'].border = THIN_BORDER
    ws[f'B{row}'].number_format = MONEY_FMT
    ws[f'B{row}'].alignment = Alignment(horizontal='right')
    for col in ['C', 'D']:
        ws[f'{col}{row}'].fill = fill
        ws[f'{col}{row}'].border = THIN_BORDER
    ws.row_dimensions[row].height = 16


def _total_row(ws, row, label, val):
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].fill = LTBLUE_FILL
        ws[f'{col}{row}'].border = MED_BORDER
    ws[f'A{row}'].value = label; ws[f'A{row}'].font = TOTAL_FONT
    ws[f'A{row}'].alignment = Alignment(indent=1)
    ws[f'B{row}'].value = val;   ws[f'B{row}'].font = TOTAL_FONT
    ws[f'B{row}'].number_format = MONEY_FMT
    ws[f'B{row}'].alignment = Alignment(horizontal='right')
    ws.row_dimensions[row].height = 18


# ── 1. TREASURER REPORT ───────────────────────────────────────────────────────

def build_treasurer(ws, qb, bank, month_label, org_name, pass_through=None):
    ws.sheet_view.showGridLines = False
    for col, w in zip(['A','B','C','D'], [32,18,18,22]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:D1'); c = ws['A1']; c.value = org_name.upper()
    c.font = Font(name='Arial', bold=True, size=16, color=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:D2'); c = ws['A2']
    c.value = f'Monthly Treasurer Report - {month_label}'
    c.font = Font(name='Arial', bold=True, size=12, color=TEAL)
    c.alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:D3'); c = ws['A3']
    c.value = f'Generated: {datetime.today().strftime("%B %d, %Y")}'
    c.font = Font(name='Arial', italic=True, size=9, color='888888')
    c.alignment = Alignment(horizontal='center')

    row = 5
    row = _sec_hdr(ws, 'INCOME', row)
    row = _col_hdrs(ws, row, ['Category', 'Amount', '', ''])
    inc_s = row
    for i, (k, v) in enumerate(qb['income'].items()):
        _data_row(ws, row, k, v, shade=(i % 2 == 1)); row += 1
    _total_row(ws, row, 'Total Income', f'=SUM(B{inc_s}:B{row-1})')
    it = row; row += 2

    row = _sec_hdr(ws, 'EXPENSES', row)
    row = _col_hdrs(ws, row, ['Category', 'Amount', '', ''])
    exp_s = row
    for i, (k, v) in enumerate(qb['expenses'].items()):
        _data_row(ws, row, k, v, shade=(i % 2 == 1)); row += 1
    _total_row(ws, row, 'Total Expenses', f'=SUM(B{exp_s}:B{row-1})')
    et = row; row += 2

    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].value = 'NET INCOME / (LOSS)'
    ws[f'A{row}'].font = Font(name='Arial', bold=True, size=11, color=WHITE)
    ws[f'A{row}'].fill = NAVY_FILL
    ws[f'A{row}'].alignment = Alignment(horizontal='left', indent=1)
    row += 1
    for col in ['A','B','C','D']:
        ws[f'{col}{row}'].fill = GOLD_FILL
        ws[f'{col}{row}'].border = MED_BORDER
    ws[f'A{row}'].value = 'Net Income (Loss)'
    ws[f'A{row}'].font = TOTAL_FONT
    ws[f'A{row}'].alignment = Alignment(indent=1)
    ws[f'B{row}'].value = f'=B{it}-B{et}'
    ws[f'B{row}'].font = TOTAL_FONT
    ws[f'B{row}'].number_format = MONEY_FMT
    ws[f'B{row}'].alignment = Alignment(horizontal='right')
    row += 2

    row = _sec_hdr(ws, 'BANK RECONCILIATION - Chase Business Checking', row)
    row = _col_hdrs(ws, row, ['Description', 'Amount', '', ''])
    for i, (lbl, amt) in enumerate([
        ('Beginning Balance',            bank['beginning_balance']),
        ('(+) Deposits & Additions',     bank['total_deposits']),
        ('(-) Checks Paid',             -bank['total_checks']),
        ('(-) Electronic Withdrawals',  -bank.get('total_withdrawals', 0.0)),
        ('(-) Fees',                    -bank['total_fees']),
    ]):
        _data_row(ws, row, lbl, amt, shade=(i % 2 == 1)); row += 1
    ending_bal_row = row
    _total_row(ws, row, 'Ending Balance (per statement)',
               bank['ending_balance']); row += 1

    calc = (bank['beginning_balance'] + bank['total_deposits']
            - bank['total_checks']
            - bank.get('total_withdrawals', 0.0)
            - bank['total_fees'])
    diff = calc - bank['ending_balance']

    for lbl, val, fill in [
        ('Calculated Ending Balance', calc, PatternFill()),
        ('Difference (should be $0.00)', diff,
         GREEN_FILL if abs(diff) < 0.01 else RED_FILL),
    ]:
        for col in ['A','B','C','D']:
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = THIN_BORDER
        ws[f'A{row}'].value = lbl; ws[f'A{row}'].font = BOLD_FONT
        ws[f'A{row}'].alignment = Alignment(indent=2)
        ws[f'B{row}'].value = val; ws[f'B{row}'].font = BOLD_FONT
        ws[f'B{row}'].number_format = MONEY_FMT
        ws[f'B{row}'].alignment = Alignment(horizontal='right')
        ws.row_dimensions[row].height = 16; row += 1
    row += 1

    if pass_through is not None:
        fund_name = pass_through['fund_name']
        row = _sec_hdr(ws, f'{fund_name} (Pass-Through Funds Held for School)', row)
        row = _col_hdrs(ws, row, ['Description', 'Amount', '', ''])
        _data_row(ws, row, f'(+) {fund_name} Deposits This Month',
                  pass_through['income_total']); row += 1
        _data_row(ws, row, f'(-) {fund_name} Payouts This Month',
                  -pass_through['expense_total'], shade=True); row += 1
        readthon_bal_row = row
        _total_row(ws, row, f'{fund_name} Balance Held in Account (Cumulative)',
                   pass_through['balance_held']); row += 2

        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'].value = 'TOTAL MONEY IN ACCOUNT / PTA MONEY'
        ws[f'A{row}'].font = Font(name='Arial', bold=True, size=11, color=WHITE)
        ws[f'A{row}'].fill = NAVY_FILL
        ws[f'A{row}'].alignment = Alignment(horizontal='left', indent=1)
        row += 1

        pta_money_val = bank['ending_balance'] - pass_through['balance_held']
        for lbl, formula, fill in [
            ('Total Money in Account', f'=B{ending_bal_row}', GOLD_FILL),
            ('PTA Money', f'=B{ending_bal_row}-B{readthon_bal_row}',
             RED_FILL if pta_money_val < 0 else GOLD_FILL),
        ]:
            for col in ['A','B','C','D']:
                ws[f'{col}{row}'].fill = fill
                ws[f'{col}{row}'].border = MED_BORDER
            ws[f'A{row}'].value = lbl; ws[f'A{row}'].font = TOTAL_FONT
            ws[f'A{row}'].alignment = Alignment(indent=1)
            ws[f'B{row}'].value = formula; ws[f'B{row}'].font = TOTAL_FONT
            ws[f'B{row}'].number_format = MONEY_FMT
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1
        row += 1

    if bank['daily_balances']:
        row = _sec_hdr(ws, 'DAILY ENDING BALANCES', row)
        for col, hdr in zip(['A','B'], ['Date', 'Balance']):
            ws[f'{col}{row}'].value = hdr
            ws[f'{col}{row}'].font = SUBHDR_FONT
            ws[f'{col}{row}'].fill = TEAL_FILL
            ws[f'{col}{row}'].border = THIN_BORDER
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
        row += 1
        for i, (dt, bal) in enumerate(sorted(bank['daily_balances'].items())):
            fill = LGREY_FILL if i % 2 else PatternFill()
            ws[f'A{row}'].value = dt;  ws[f'A{row}'].font = BODY_FONT
            ws[f'A{row}'].fill = fill; ws[f'A{row}'].border = THIN_BORDER
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            ws[f'B{row}'].value = bal; ws[f'B{row}'].font = BODY_FONT
            ws[f'B{row}'].fill = fill; ws[f'B{row}'].border = THIN_BORDER
            ws[f'B{row}'].number_format = MONEY_FMT
            ws[f'B{row}'].alignment = Alignment(horizontal='right')
            row += 1


# ── 2. BUDGET VS ACTUALS ──────────────────────────────────────────────────────

def build_budget(ws, title, merged_data, org_name, fiscal_months,
                 current_idx, fiscal_year_start, show_pl=False, income_merged=None):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 13
    ws.column_dimensions['C'].width = 13
    for i in range(12):
        ws.column_dimensions[get_column_letter(4+i)].width = 9
    ws.column_dimensions[get_column_letter(16)].width = 11
    ws.column_dimensions[get_column_letter(17)].width = 11

    ws.merge_cells(f'A1:{get_column_letter(17)}1'); c = ws['A1']
    c.value = f'{org_name.upper()}  -  {title}'
    c.font = Font(name='Arial', bold=True, size=13, color=WHITE)
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f'A2:{get_column_letter(17)}2'); c = ws['A2']
    c.value = (f'As of {datetime.today().strftime("%B %d, %Y")}  |  '
               f'Fiscal Year {fiscal_year_start} - {fiscal_year_start + 1}  |  '
               f'Active month: {fiscal_months[current_idx]}')
    c.font = Font(name='Arial', italic=True, size=9, color='666666')
    c.alignment = Alignment(horizontal='center')

    pl_label = 'Profit/Loss' if show_pl else ''
    headers = (['Category', 'Last Year', 'Budget (Annual)']
               + fiscal_months + ['Total', pl_label])
    for ci, hdr in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=hdr)
        is_active = (4 <= ci <= 15 and (ci-4) == current_idx)
        c.font = Font(name='Arial', bold=True, size=9,
                      color=NAVY if is_active else WHITE)
        c.fill = GOLD_FILL if is_active else TEAL_FILL
        c.alignment = Alignment(
            horizontal='center' if ci > 1 else 'left',
            vertical='center', wrap_text=True)
        c.border = THIN_BORDER
    ws.row_dimensions[3].height = 30

    # Build income lookup for P&L column
    inc_lookup = {}
    if show_pl and income_merged:
        inc_lookup = {
            item.lower(): sum(monthly[:current_idx+1])
            for section in income_merged.values()
            for item, (_, _, monthly) in section.items()
        }

    dr = 4
    for section, items in merged_data.items():
        ws.merge_cells(f'A{dr}:{get_column_letter(17)}{dr}')
        c = ws[f'A{dr}']; c.value = section
        c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[dr].height = 18
        ss = dr + 1; dr += 1

        for ir, (item, (last_yr, budget, monthly)) in enumerate(items.items()):
            fill = LGREY_FILL if ir % 2 == 1 else PatternFill()
            c = ws.cell(row=dr, column=1, value=item)
            c.font = BODY_FONT; c.fill = fill
            c.alignment = Alignment(indent=2); c.border = THIN_BORDER

            for ci, v in [(2, last_yr), (3, budget)]:
                c = ws.cell(row=dr, column=ci, value=v if v else None)
                c.font = BODY_FONT; c.fill = fill
                c.number_format = MONEY_FMT
                c.alignment = Alignment(horizontal='right')
                c.border = THIN_BORDER

            for mi, v in enumerate(monthly):
                is_active = (mi == current_idx)
                c = ws.cell(row=dr, column=4+mi, value=v if v else None)
                c.font = Font(name='Arial', bold=is_active, size=10)
                c.fill = (GOLD_FILL if (is_active and v)
                          else PatternFill('solid', fgColor='FFF9E6')
                          if is_active else fill)
                c.number_format = MONEY_FMT
                c.alignment = Alignment(horizontal='right')
                c.border = THIN_BORDER

            c = ws.cell(row=dr, column=16, value=f'=SUM(D{dr}:O{dr})')
            c.font = BODY_FONT; c.fill = fill; c.number_format = MONEY_FMT
            c.alignment = Alignment(horizontal='right'); c.border = THIN_BORDER

            # Profit/Loss column
            if show_pl:
                act_exp = sum(monthly[:current_idx+1])
                act_inc = inc_lookup.get(item.lower(), 0.0)
                pl_val  = (act_inc - act_exp
                           if (act_inc or act_exp) else None)
            else:
                pl_val = None
            c = ws.cell(row=dr, column=17, value=pl_val)
            c.font = BODY_FONT; c.fill = fill; c.number_format = MONEY_FMT
            c.alignment = Alignment(horizontal='right'); c.border = THIN_BORDER

            ws.row_dimensions[dr].height = 15; dr += 1

        se = dr - 1
        c = ws.cell(row=dr, column=1, value=f'Total {section}')
        c.font = TOTAL_FONT; c.fill = LTBLUE_FILL
        c.alignment = Alignment(indent=1); c.border = MED_BORDER
        for ci in range(2, 18):
            cl = get_column_letter(ci)
            c = ws.cell(row=dr, column=ci,
                        value=f'=SUM({cl}{ss}:{cl}{se})')
            c.font = TOTAL_FONT; c.fill = LTBLUE_FILL
            c.number_format = MONEY_FMT
            c.alignment = Alignment(horizontal='right'); c.border = MED_BORDER
        ws.row_dimensions[dr].height = 18; dr += 2

    c = ws.cell(row=dr, column=1, value='GRAND TOTAL')
    c.font = Font(name='Arial', bold=True, size=11, color=WHITE)
    c.fill = NAVY_FILL; c.alignment = Alignment(indent=1); c.border = MED_BORDER
    for ci in range(2, 18):
        cl = get_column_letter(ci)
        c = ws.cell(row=dr, column=ci,
                    value=f'=SUMIF(A4:A{dr-1},"Total*",{cl}4:{cl}{dr-1})')
        c.font = Font(name='Arial', bold=True, size=10, color=WHITE)
        c.fill = NAVY_FILL; c.number_format = MONEY_FMT
        c.alignment = Alignment(horizontal='right'); c.border = MED_BORDER
    ws.row_dimensions[dr].height = 20


# ── 3. GIVEBACK RECONCILIATION ────────────────────────────────────────────────

def build_givebacks(ws, givebacks, bank, org_name):
    ws.sheet_view.showGridLines = False
    for col, w in zip(['A','B','C','D','E','F'], [30,18,12,16,11,25]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:F1'); c = ws['A1']
    c.value = f'{org_name.upper()}  -  Giveback Reconciliation'
    c.font = Font(name='Arial', bold=True, size=13, color=WHITE)
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:F2'); c = ws['A2']
    c.value = f'Generated: {datetime.today().strftime("%B %d, %Y")}'
    c.font = Font(name='Arial', italic=True, size=9, color='666666')
    c.alignment = Alignment(horizontal='center')

    row = 4
    for col, hdr in zip(['A','B','C','D','E','F'],
                         ['Item','Category','Transactions',
                          'Amount','% of Total','Source File']):
        c = ws[f'{col}{row}']; c.value = hdr; c.font = SUBHDR_FONT
        c.fill = TEAL_FILL
        c.alignment = Alignment(
            horizontal='center' if col != 'A' else 'left')
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 18; row += 1

    ds = row; trn = ds + len(givebacks)
    for i, g in enumerate(givebacks):
        fill = LGREY_FILL if i % 2 == 1 else PatternFill()
        for col in ['A','B','C','D','E','F']:
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = THIN_BORDER
        ws[f'A{row}'].value = g['item']
        ws[f'A{row}'].font = BODY_FONT
        ws[f'A{row}'].alignment = Alignment(indent=1)
        ws[f'B{row}'].value = g['category']
        ws[f'B{row}'].font = BODY_FONT
        ws[f'B{row}'].alignment = Alignment(horizontal='center')
        ws[f'C{row}'].value = g['count']
        ws[f'C{row}'].font = BODY_FONT
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        ws[f'D{row}'].value = g['total']
        ws[f'D{row}'].font = BODY_FONT
        ws[f'D{row}'].number_format = MONEY_FMT
        ws[f'D{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].value = f'=D{row}/D{trn}'
        ws[f'E{row}'].font = BODY_FONT
        ws[f'E{row}'].number_format = '0.0%'
        ws[f'E{row}'].alignment = Alignment(horizontal='center')
        ws[f'F{row}'].value = g.get('source_file', '')
        ws[f'F{row}'].font = BODY_FONT
        ws[f'F{row}'].alignment = Alignment(indent=1)
        ws.row_dimensions[row].height = 15; row += 1

    for col in ['A','B','C','D','E','F']:
        ws[f'{col}{row}'].fill = LTBLUE_FILL
        ws[f'{col}{row}'].border = MED_BORDER
    ws[f'A{row}'].value = 'TOTAL'; ws[f'A{row}'].font = TOTAL_FONT
    ws[f'A{row}'].alignment = Alignment(indent=1)
    ws[f'C{row}'].value = f'=SUM(C{ds}:C{row-1})'
    ws[f'C{row}'].font = TOTAL_FONT
    ws[f'C{row}'].alignment = Alignment(horizontal='center')
    ws[f'D{row}'].value = f'=SUM(D{ds}:D{row-1})'
    ws[f'D{row}'].font = TOTAL_FONT
    ws[f'D{row}'].number_format = MONEY_FMT
    ws[f'D{row}'].alignment = Alignment(horizontal='right')
    ws[f'E{row}'].value = '100.0%'; ws[f'E{row}'].font = TOTAL_FONT
    ws[f'E{row}'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 18; row += 2

    _sec_hdr(ws, 'GIVEBACK <-> BANK RECONCILIATION', row, n=6); row += 1
    gb_total = sum(g['total'] for g in givebacks)
    # A month can have more than one GB Payout deposit clear separately
    # (e.g. two payouts landing days apart) -- sum all matches, not just
    # the first one, or a real second deposit gets silently dropped and
    # shows as a false unreconciled "Difference".
    bank_gb  = sum(
        d['amount'] for d in bank['deposits']
        if 'gb payout' in d.get('description','').lower()
        or 'givebacks' in d.get('description','').lower())

    for lbl, val in [
        ('Givebacks Platform Total',            gb_total),
        ('Givebacks Deposit in Bank Statement',  bank_gb),
        ('Difference',                           gb_total - bank_gb),
    ]:
        is_d = lbl == 'Difference'
        fill = (GREEN_FILL if abs(gb_total-bank_gb) < 0.01
                else RED_FILL) if is_d else PatternFill()
        for col in ['A','B','C','D','E','F']:
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = THIN_BORDER
        ws[f'A{row}'].value = lbl; ws[f'A{row}'].font = BOLD_FONT
        ws[f'A{row}'].alignment = Alignment(indent=2)
        ws[f'D{row}'].value = val; ws[f'D{row}'].font = BOLD_FONT
        ws[f'D{row}'].number_format = MONEY_FMT
        ws[f'D{row}'].alignment = Alignment(horizontal='right')
        ws.row_dimensions[row].height = 16; row += 1


# ── 4. FILE MANIFEST ──────────────────────────────────────────────────────────

def build_manifest(ws, month_folder, org_name, month_label,
                   fiscal_idx, fiscal_months):
    ws.sheet_view.showGridLines = False
    for col, w in zip(['A','B','C','D'], [15,35,22,14]):
        ws.column_dimensions[col].width = w

    ws.merge_cells('A1:D1'); c = ws['A1']
    c.value = f'{org_name.upper()}  -  File Manifest'
    c.font = Font(name='Arial', bold=True, size=13, color=WHITE)
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:D2'); c = ws['A2']
    c.value = (f'Report: {month_label}  |  '
               f'Column: {fiscal_months[fiscal_idx]}  |  '
               f'Generated: {datetime.today().strftime("%B %d, %Y at %I:%M %p")}')
    c.font = Font(name='Arial', italic=True, size=9, color='666666')
    c.alignment = Alignment(horizontal='center')

    row = 4
    for col, hdr in zip(['A','B','C','D'],
                         ['Type','Filename','Last Modified','Size']):
        c = ws[f'{col}{row}']; c.value = hdr; c.font = SUBHDR_FONT
        c.fill = TEAL_FILL
        c.alignment = Alignment(
            horizontal='left' if col == 'B' else 'center')
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = 18; row += 1

    month_folder = Path(month_folder)
    all_files = (
        [('QuickBooks', f) for f in sorted(month_folder.glob('*.csv'))
         if 'givebacks' not in f.name.lower()] +
        [('Bank Statement', f) for f in sorted(month_folder.glob('*.pdf'))] +
        [('Givebacks', f) for f in sorted(
            (month_folder / 'givebacks').glob('*.csv'))
         if (month_folder / 'givebacks').exists()]
    )

    for i, (ft, fp) in enumerate(all_files):
        fill = LGREY_FILL if i % 2 == 1 else PatternFill()
        stat = fp.stat()
        mod  = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M')
        size = f'{stat.st_size/1024:.1f} KB'
        for col in ['A','B','C','D']:
            ws[f'{col}{row}'].fill = fill
            ws[f'{col}{row}'].border = THIN_BORDER
        ws[f'A{row}'].value = ft; ws[f'A{row}'].font = BODY_FONT
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        ws[f'B{row}'].value = fp.name; ws[f'B{row}'].font = BODY_FONT
        ws[f'B{row}'].alignment = Alignment(indent=1)
        ws[f'C{row}'].value = mod; ws[f'C{row}'].font = BODY_FONT
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        ws[f'D{row}'].value = size; ws[f'D{row}'].font = BODY_FONT
        ws[f'D{row}'].alignment = Alignment(horizontal='right', indent=1)
        ws.row_dimensions[row].height = 16; row += 1

    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'].value = f'Total files processed: {len(all_files)}'
    ws[f'A{row}'].font = BOLD_FONT
    ws[f'A{row}'].alignment = Alignment(indent=1)


# ── 5. YTD SUMMARY ───────────────────────────────────────────────────────────
def build_ytd_summary(ws, income_merged, expense_merged, org_name,
                      month_label, fiscal_idx, fiscal_months, fiscal_year_start,
                      bank, balance_forward=0.0, pass_through=None):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 28
    for col in ['B','C','D','E','F','G','H','I']:
        ws.column_dimensions[col].width = 14

    # Title
    ws.merge_cells('A1:I1'); c = ws['A1']
    c.value = f"{org_name}  —  Treasurer's Report"
    c.font = Font(name='Arial', bold=True, size=14, color=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:I2'); c = ws['A2']
    c.value = (f'7/1/{fiscal_year_start} - '
               f'{datetime.today().strftime("%-m/%-d/%Y")}')
    c.font = Font(name='Arial', italic=True, size=10, color='666666')
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16

    # Balance forward and current balance
    ws['A3'].value = 'Balance Forward:'
    ws['A3'].font = BOLD_FONT
    ws['B3'].value = balance_forward
    ws['B3'].font = BOLD_FONT
    ws['B3'].number_format = MONEY_FMT
    ws['C3'].value = f'as of 6/30/{fiscal_year_start + 1}'
    ws['C3'].font = Font(name='Arial', italic=True, size=9, color='666666')

    ws['E3'].value = 'Current Balance:'
    ws['E3'].font = BOLD_FONT
    ws['F3'].value = bank['ending_balance']
    ws['F3'].font = BOLD_FONT
    ws['F3'].number_format = MONEY_FMT
    period_end = (bank.get('period','').split('through')[-1].strip()
                  if bank.get('period') else '')
    ws['G3'].value = f'as of {period_end}'
    ws['G3'].font = Font(name='Arial', italic=True, size=9, color='666666')

    # PTA Money = current balance minus pass-through fund money still held
    # -- only shown when there's actually some held (nothing to subtract
    # otherwise, so it would just duplicate Current Balance).
    if pass_through is not None and pass_through.get('balance_held'):
        ws['H3'].value = 'PTA Money:'
        ws['H3'].font = BOLD_FONT
        ws['I3'].value = bank['ending_balance'] - pass_through['balance_held']
        ws['I3'].font = BOLD_FONT
        ws['I3'].number_format = MONEY_FMT

    ws.row_dimensions[3].height = 18

    # YTD totals
    ytd_inc = sum(
        sum(monthly[:fiscal_idx+1])
        for section in income_merged.values()
        for _, (_, _, monthly) in section.items()
    )
    ytd_exp = sum(
        sum(monthly[:fiscal_idx+1])
        for section in expense_merged.values()
        for _, (_, _, monthly) in section.items()
    )
    ws['A4'].value = f'Total Income: ${ytd_inc:,.2f}'
    ws['A4'].font = BOLD_FONT
    ws['E4'].value = f'Total Expenses: ${ytd_exp:,.2f}'
    ws['E4'].font = BOLD_FONT
    ws.row_dimensions[4].height = 18

    # Build lookups
    income_lookup = {}
    for section, items in income_merged.items():
        for item, (last_yr, budget, monthly) in items.items():
            income_lookup[item.lower()] = (last_yr, budget, monthly)

    expense_lookup = {}
    for section, items in expense_merged.items():
        for item, (last_yr, budget, monthly) in items.items():
            expense_lookup[item.lower()] = (last_yr, budget, monthly)

    # Build section structure from expense sections
    section_items = {}
    for section, items in expense_merged.items():
        section_items[section] = {}
        for item, (last_yr_exp, exp_budget, exp_monthly) in items.items():
            inc_data    = income_lookup.get(item.lower())
            last_yr_inc = inc_data[0] if inc_data else 0.0
            inc_budget  = inc_data[1] if inc_data else 0.0
            inc_monthly = inc_data[2] if inc_data else [0.0]*12
            section_items[section][item] = {
                'last_yr_inc':  last_yr_inc,
                'last_yr_exp':  last_yr_exp,
                'exp_budget':   exp_budget,
                'inc_budget':   inc_budget,
                'act_income':   sum(inc_monthly[:fiscal_idx+1]),
                'act_expenses': sum(exp_monthly[:fiscal_idx+1]),
            }

    # Add income-only items
    # Special merged rows
    MERGE_PAIRS = {
        'membership expenses': {
            'label':        'Membership Expenses/Income',
            'income_keys':  ['single', 'family', 'donations', 'teachers', 'student'],
            'expense_keys': ['membership expenses', 'council dues'],
        },
        'entertainment': {
            'label':        'Basket Dinner',
            'income_keys':  ['ticket & raffle sales', 'sponsors'],
            'expense_keys': ['entertainment', 'raffles', 'venue'],
        },
    }
    
    all_merged_income_keys  = set()
    all_merged_expense_keys = set()
    for merge in MERGE_PAIRS.values():
        all_merged_income_keys.update(merge['income_keys'])
        all_merged_expense_keys.update(merge.get('expense_keys', []))
    
    other_income = {}
    
    for section, items in income_merged.items():
        for item, (last_yr_inc, inc_budget, inc_monthly) in items.items():
            if item.lower() not in expense_lookup and item.lower() not in all_merged_income_keys:
                other_income[item] = {
                    'last_yr_inc':  last_yr_inc,
                    'last_yr_exp':  0.0,
                    'exp_budget':   0.0,
                    'inc_budget':   inc_budget,
                    'act_income':   sum(inc_monthly[:fiscal_idx+1]),
                    'act_expenses': 0.0,
                }
    if other_income:
        section_items['Other Income'] = other_income

    def money_or_none(val):
        return val if val and abs(val) > 0.001 else None

    def money_or_dash(val):       # ← add this
        if val is None or abs(val) < 0.001:
            return '--'
        return val

    def sec_hdr_ytd(label, r):
        ws.merge_cells(f'A{r}:I{r}')
        c = ws[f'A{r}']; c.value = label
        c.font = Font(name='Arial', bold=True, size=11, color=WHITE)
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[r].height = 20
        return r + 1

    def col_hdrs_ytd(r):
        for ci, hdr in enumerate([
            'Category',
            "Last Year's Income", "Last Year's Expenses",
            'Expected Income',    'Expected Expenses',
            'Actual Income',      'Actual Expenses',
            'Net Profit',         "Last Year's Profit",
        ], 1):
            c = ws.cell(row=r, column=ci, value=hdr)
            c.font = Font(name='Arial', bold=True, size=8, color=WHITE)
            c.fill = TEAL_FILL; c.border = THIN_BORDER
            c.alignment = Alignment(
                horizontal='center' if ci > 1 else 'left',
                vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 28
        return r + 1

    row = 6
    row = col_hdrs_ytd(row)
    
    for section_name, items in section_items.items():
        if not items:
            continue

        row = sec_hdr_ytd(section_name, row)
        #row = col_hdrs_ytd(row)

        sec = defaultdict(float)

        rendered_income_keys = set()  # track which income items were merged
        rendered_expense_keys = set()
        
        for i, (item, vals) in enumerate(items.items()):
            item_lower = item.lower()

            # Skip items already merged into another row
            if (item_lower in all_merged_expense_keys and
                    item_lower not in MERGE_PAIRS):
                continue

            if item_lower in MERGE_PAIRS:
                merge = MERGE_PAIRS[item_lower]
                
                # Sum income
                merged_inc_actual  = sum(
                    sum(income_lookup[k][2][:fiscal_idx+1])
                    for k in merge['income_keys'] if k in income_lookup)
                merged_inc_last_yr = sum(
                    income_lookup[k][0]
                    for k in merge['income_keys'] if k in income_lookup)
                merged_inc_budget  = sum(
                    income_lookup[k][1]
                    for k in merge['income_keys'] if k in income_lookup)

                # Sum expenses
                merged_exp_actual  = sum(
                    sum(expense_lookup[k][2][:fiscal_idx+1])
                    for k in merge.get('expense_keys', []) if k in expense_lookup)
                merged_exp_last_yr = sum(
                    expense_lookup[k][0]
                    for k in merge.get('expense_keys', []) if k in expense_lookup)
                merged_exp_budget  = sum(
                    expense_lookup[k][1]
                    for k in merge.get('expense_keys', []) if k in expense_lookup)

                rendered_income_keys.update(merge['income_keys'])
                rendered_expense_keys.update(merge.get('expense_keys', []))

                display_vals = {
                    'last_yr_inc':  merged_inc_last_yr,
                    'last_yr_exp':  merged_exp_last_yr,
                    'inc_budget':   merged_inc_budget,
                    'exp_budget':   merged_exp_budget,
                    'act_income':   merged_inc_actual,
                    'act_expenses': merged_exp_actual,
                }
                display_label = merge['label']
            else:
                display_label = item
                display_vals  = vals

            has_data = (
                abs(display_vals['last_yr_inc'])  > 0.001 or
                abs(display_vals['last_yr_exp'])  > 0.001 or
                abs(display_vals['inc_budget'])   > 0.001 or
                abs(display_vals['exp_budget'])   > 0.001 or
                abs(display_vals['act_income'])   > 0.001 or
                abs(display_vals['act_expenses']) > 0.001
            )
            if not has_data:
                continue

            fill = LGREY_FILL if i % 2 == 1 else PatternFill()
            net      = display_vals['act_income'] - display_vals['act_expenses']
            last_net = display_vals['last_yr_inc'] - display_vals['last_yr_exp']

            row_vals = [
                display_label,
                money_or_dash(display_vals['last_yr_inc']),
                money_or_dash(display_vals['last_yr_exp']),
                money_or_dash(display_vals['inc_budget']),
                money_or_dash(display_vals['exp_budget']),
                money_or_dash(display_vals['act_income']),
                money_or_dash(display_vals['act_expenses']),
                money_or_dash(net),
                money_or_dash(last_net),
            ]

            # ... rest of cell rendering stays the same
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.font = BODY_FONT; c.fill = fill; c.border = THIN_BORDER
                if ci == 1:
                    c.alignment = Alignment(indent=2)
                else:
                    c.number_format = MONEY_FMT
                    c.alignment = Alignment(horizontal='right')

            # Highlight net profit
            if vals['act_income'] > 0 or vals['act_expenses'] > 0:
                net_cell = ws.cell(row=row, column=8)
                net_cell.fill = GREEN_FILL if net >= 0 else RED_FILL

            ws.row_dimensions[row].height = 15

            for k in ['last_yr_inc','last_yr_exp','inc_budget',
                      'exp_budget','act_income','act_expenses']:
                sec[k] += display_vals[k]
            row += 1

        # Section total
        sec_net      = sec['act_income'] - sec['act_expenses']
        sec_last_net = sec['last_yr_inc'] - sec['last_yr_exp']

        for ci in range(1, 10):
            ws.cell(row=row, column=ci).fill   = LTBLUE_FILL
            ws.cell(row=row, column=ci).border = MED_BORDER

        ws.cell(row=row, column=1, value='TOTAL:').font = TOTAL_FONT
        ws.cell(row=row, column=1).alignment = Alignment(indent=1)

        for ci, val in enumerate([
            money_or_dash(sec['last_yr_inc']),
            money_or_dash(sec['last_yr_exp']),
            money_or_dash(sec['inc_budget']),
            money_or_dash(sec['exp_budget']),
            money_or_dash(sec['act_income']),
            money_or_dash(sec['act_expenses']),
            money_or_dash(sec_net),
            money_or_dash(sec_last_net),
        ], 2):
            c = ws.cell(row=row, column=ci, value=val)
            c.font = TOTAL_FONT; c.number_format = MONEY_FMT
            c.alignment = Alignment(horizontal='right')

        ws.row_dimensions[row].height = 18
        row += 2


def build_ytd_summary_compact(ws, income_merged, expense_merged, org_name,
                               month_label, fiscal_idx, fiscal_months,
                               fiscal_year_start, bank, balance_forward=0.0,
                               pass_through=None):
    """
    Compact YTD Summary: 5 columns (Category, Budget, Actual, Last Year,
    Profit/Loss) instead of build_ytd_summary's 9. A category with both
    an income and an expense side (e.g. Book Fair) renders as two rows --
    "<item> — Income" / "<item> — Expense" -- rather than cramming two
    numbers into one cell, so every Budget/Actual/Last Year cell stays a
    real, summable number (a single cell can't hold two numbers and still
    work in an Excel SUM() the way a treasurer select-and-totaling a
    column would expect). Section TOTAL rows are the one exception --
    those are synthetic aggregates already, so their Budget/Actual/Last
    Year cells show combined "$income / $expense" text since nothing
    above them depends on summing a TOTAL row further.

    build_ytd_summary (the original 9-column version) is kept as-is,
    still fully working, for on-demand use whenever the regular version
    is asked for specifically -- not called by the default pipeline, but
    not deleted either.

    balance_forward: caller is expected to have already subtracted the
    pass-through balance held AS OF the fiscal year's own July 1 (not
    today's cumulative figure) -- see pipeline.get_fiscal_year_balance_forward,
    which already does this. Current Balance is computed here directly,
    since this function already has both bank and pass_through in hand.
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 32
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 18

    ws.merge_cells('A1:E1'); c = ws['A1']
    c.value = f"{org_name}  —  Treasurer's Report"
    c.font = Font(name='Arial', bold=True, size=14, color=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:E2'); c = ws['A2']
    c.value = (f'7/1/{fiscal_year_start} - '
               f'{datetime.today().strftime("%-m/%-d/%Y")}')
    c.font = Font(name='Arial', italic=True, size=10, color='666666')
    c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 16

    # Both figures reflect PTA money -- net of any pass-through fund
    # balance held in the account -- rather than the raw bank balance, so
    # this row can't be misread as "how much cash is in the account" when
    # part of that cash isn't really the org's.
    pass_through_balance_held = pass_through['balance_held'] if pass_through else 0.0
    current_balance = bank['ending_balance'] - pass_through_balance_held

    ws['A3'].value = 'Balance Forward (PTA):'
    ws['A3'].font = BOLD_FONT
    ws['B3'].value = balance_forward
    ws['B3'].font = BOLD_FONT
    ws['B3'].number_format = MONEY_FMT

    ws['D3'].value = 'Current Balance (PTA):'
    ws['D3'].font = BOLD_FONT
    ws['E3'].value = current_balance
    ws['E3'].font = BOLD_FONT
    ws['E3'].number_format = MONEY_FMT
    ws.row_dimensions[3].height = 18

    ytd_inc = sum(sum(monthly[:fiscal_idx+1])
                  for section in income_merged.values()
                  for _, (_, _, monthly) in section.items())
    ytd_exp = sum(sum(monthly[:fiscal_idx+1])
                  for section in expense_merged.values()
                  for _, (_, _, monthly) in section.items())
    ws['A4'].value = (f'Total Income: ${ytd_inc:,.2f}   |   '
                       f'Total Expenses: ${ytd_exp:,.2f}   |   '
                       f'Net: ${ytd_inc - ytd_exp:,.2f}')
    ws['A4'].font = BOLD_FONT
    ws.row_dimensions[4].height = 18

    income_lookup = {}
    for section, items in income_merged.items():
        for item, (last_yr, budget, monthly) in items.items():
            income_lookup[item.lower()] = (last_yr, budget, monthly)

    expense_lookup = {}
    for section, items in expense_merged.items():
        for item, (last_yr, budget, monthly) in items.items():
            expense_lookup[item.lower()] = (last_yr, budget, monthly)

    section_items = {}
    for section, items in expense_merged.items():
        section_items[section] = {}
        for item, (last_yr_exp, exp_budget, exp_monthly) in items.items():
            inc_data    = income_lookup.get(item.lower())
            last_yr_inc = inc_data[0] if inc_data else 0.0
            inc_budget  = inc_data[1] if inc_data else 0.0
            inc_monthly = inc_data[2] if inc_data else [0.0]*12
            section_items[section][item] = {
                'last_yr_inc':  last_yr_inc,
                'last_yr_exp':  last_yr_exp,
                'exp_budget':   exp_budget,
                'inc_budget':   inc_budget,
                'act_income':   sum(inc_monthly[:fiscal_idx+1]),
                'act_expenses': sum(exp_monthly[:fiscal_idx+1]),
            }

    MERGE_PAIRS = {
        'membership expenses': {
            # Just "Membership" here (unlike build_ytd_summary's fuller
            # "Membership Expenses/Income") -- the compact sheet already
            # splits this into separate Income/Expense rows, so repeating
            # both words in the label is redundant.
            'label':        'Membership',
            'income_keys':  ['single', 'family', 'donations', 'teachers', 'student'],
            'expense_keys': ['membership expenses', 'council dues'],
        },
        'entertainment': {
            'label':        'Basket Dinner',
            'income_keys':  ['ticket & raffle sales', 'sponsors'],
            'expense_keys': ['entertainment', 'raffles', 'venue'],
        },
    }
    all_merged_income_keys  = set()
    all_merged_expense_keys = set()
    for merge in MERGE_PAIRS.values():
        all_merged_income_keys.update(merge['income_keys'])
        all_merged_expense_keys.update(merge.get('expense_keys', []))

    other_income = {}
    for section, items in income_merged.items():
        for item, (last_yr_inc, inc_budget, inc_monthly) in items.items():
            if item.lower() not in expense_lookup and item.lower() not in all_merged_income_keys:
                other_income[item] = {
                    'last_yr_inc':  last_yr_inc,
                    'last_yr_exp':  0.0,
                    'exp_budget':   0.0,
                    'inc_budget':   inc_budget,
                    'act_income':   sum(inc_monthly[:fiscal_idx+1]),
                    'act_expenses': 0.0,
                }
    if other_income:
        section_items['Other Income'] = other_income

    def has_val(*vals):
        return any(abs(v) > 0.001 for v in vals)

    def money_or_dash(val):
        if val is None or abs(val) < 0.001:
            return '--'
        return val

    def combo_text(inc, exp):
        if abs(inc) < 0.001 and abs(exp) < 0.001:
            return '--'
        if abs(inc) < 0.001:
            return f'${exp:,.2f}'
        if abs(exp) < 0.001:
            return f'${inc:,.2f}'
        return f'${inc:,.2f} / ${exp:,.2f}'

    def sec_hdr(label, r):
        ws.merge_cells(f'A{r}:E{r}')
        c = ws[f'A{r}']; c.value = label
        c.font = Font(name='Arial', bold=True, size=11, color=WHITE)
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[r].height = 20
        return r + 1

    def col_hdrs(r):
        for ci, hdr in enumerate([
            'Category', 'Budget (Income/Expense)', 'Actual (Income/Expense)',
            'Last Year (Income/Expense)', 'Profit/Loss',
        ], 1):
            c = ws.cell(row=r, column=ci, value=hdr)
            c.font = Font(name='Arial', bold=True, size=9, color=WHITE)
            c.fill = TEAL_FILL; c.border = THIN_BORDER
            c.alignment = Alignment(horizontal='center' if ci > 1 else 'left',
                                     vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 26
        return r + 1

    # A dual-sided item (e.g. Book Fair) renders as two rows -- this pair
    # of borders draws a bracket around them (medium navy on the outer
    # edges, thin on the seam between the two), so the pair visually
    # reads as one unit without merging cells or losing per-row numbers.
    PAIR_TOP_BORDER    = Border(left=MED, right=MED, top=MED, bottom=THIN)
    PAIR_BOTTOM_BORDER = Border(left=MED, right=MED, top=THIN, bottom=MED)
    SUBLABEL_FONT  = Font(name='Arial', size=10, italic=True, color='666666')
    # Last Year is reference/historical, not this year's actionable
    # numbers -- greyed + italic so the eye lands on Budget/Actual/
    # Profit-Loss first.
    LASTYEAR_FONT       = Font(name='Arial', size=10, italic=True, color='808080')
    LASTYEAR_TOTAL_FONT = Font(name='Arial', size=10, italic=True, color='808080', bold=True)

    def render_row(r, fill, label, budget, actual, last_yr, profit_loss,
                    highlight=None, border=THIN_BORDER, label_font=None, indent=2):
        for ci, val in enumerate(
                [label, money_or_dash(budget), money_or_dash(actual),
                 money_or_dash(last_yr), money_or_dash(profit_loss)], 1):
            c = ws.cell(row=r, column=ci, value=val)
            if ci == 1:
                c.font = label_font if label_font else BODY_FONT
            elif ci == 4:
                c.font = LASTYEAR_FONT
            else:
                c.font = BODY_FONT
            c.fill = fill; c.border = border
            if ci == 1:
                c.alignment = Alignment(indent=indent)
            else:
                c.number_format = MONEY_FMT
                c.alignment = Alignment(horizontal='right')
        if highlight is not None:
            ws.cell(row=r, column=5).fill = GREEN_FILL if highlight >= 0 else RED_FILL

    row = 6
    row = col_hdrs(row)

    for section_name, items in section_items.items():
        if not items:
            continue
        row = sec_hdr(section_name, row)
        sec = defaultdict(float)
        i = 0

        for item, vals in items.items():
            item_lower = item.lower()
            if item_lower in all_merged_expense_keys and item_lower not in MERGE_PAIRS:
                continue

            if item_lower in MERGE_PAIRS:
                merge = MERGE_PAIRS[item_lower]
                display_label = merge['label']
                display_vals = {
                    'last_yr_inc': sum(income_lookup[k][0] for k in merge['income_keys'] if k in income_lookup),
                    'last_yr_exp': sum(expense_lookup[k][0] for k in merge.get('expense_keys', []) if k in expense_lookup),
                    'inc_budget':  sum(income_lookup[k][1] for k in merge['income_keys'] if k in income_lookup),
                    'exp_budget':  sum(expense_lookup[k][1] for k in merge.get('expense_keys', []) if k in expense_lookup),
                    'act_income':  sum(sum(income_lookup[k][2][:fiscal_idx+1]) for k in merge['income_keys'] if k in income_lookup),
                    'act_expenses': sum(sum(expense_lookup[k][2][:fiscal_idx+1]) for k in merge.get('expense_keys', []) if k in expense_lookup),
                }
            else:
                display_label = item
                display_vals  = vals

            has_income  = has_val(display_vals['last_yr_inc'], display_vals['inc_budget'], display_vals['act_income'])
            has_expense = has_val(display_vals['last_yr_exp'], display_vals['exp_budget'], display_vals['act_expenses'])
            if not has_income and not has_expense:
                continue

            net = display_vals['act_income'] - display_vals['act_expenses']
            fill = LGREY_FILL if i % 2 == 1 else PatternFill()
            has_actuals = bool(display_vals['act_income'] or display_vals['act_expenses'])

            if has_income and has_expense:
                render_row(row, fill, f'{display_label} (Income)',
                           display_vals['inc_budget'], display_vals['act_income'], display_vals['last_yr_inc'], None,
                           border=PAIR_TOP_BORDER)
                row += 1
                render_row(row, fill, 'Expense',
                           display_vals['exp_budget'], display_vals['act_expenses'], display_vals['last_yr_exp'],
                           net, highlight=net if has_actuals else None,
                           border=PAIR_BOTTOM_BORDER, label_font=SUBLABEL_FONT, indent=4)
                row += 1
            elif has_income:
                render_row(row, fill, f'{display_label} (Income)',
                           display_vals['inc_budget'], display_vals['act_income'], display_vals['last_yr_inc'],
                           net, highlight=net if has_actuals else None)
                row += 1
            else:
                render_row(row, fill, display_label,
                           display_vals['exp_budget'], display_vals['act_expenses'], display_vals['last_yr_exp'],
                           net, highlight=net if has_actuals else None)
                row += 1

            i += 1
            for k in ['last_yr_inc', 'last_yr_exp', 'inc_budget', 'exp_budget', 'act_income', 'act_expenses']:
                sec[k] += display_vals[k]

        sec_net = sec['act_income'] - sec['act_expenses']

        for ci in range(1, 6):
            ws.cell(row=row, column=ci).fill = LTBLUE_FILL
            ws.cell(row=row, column=ci).border = MED_BORDER
        ws.cell(row=row, column=1, value='TOTAL:').font = TOTAL_FONT
        ws.cell(row=row, column=1).alignment = Alignment(indent=1)
        for ci, val in [
            (2, combo_text(sec['inc_budget'], sec['exp_budget'])),
            (3, combo_text(sec['act_income'], sec['act_expenses'])),
            (4, combo_text(sec['last_yr_inc'], sec['last_yr_exp'])),
            (5, money_or_dash(sec_net)),
        ]:
            c = ws.cell(row=row, column=ci, value=val)
            c.font = LASTYEAR_TOTAL_FONT if ci == 4 else TOTAL_FONT
            if ci == 5:
                c.number_format = MONEY_FMT
                if sec['act_income'] or sec['act_expenses']:
                    c.fill = GREEN_FILL if sec_net >= 0 else RED_FILL
            c.alignment = Alignment(horizontal='right')
        ws.row_dimensions[row].height = 18
        row += 2


# ── 6. DEBITS & CREDITS (WHOLE-FISCAL-YEAR LEDGER) ────────────────────────────
# Shared helpers for wide (5-9 column), month-sectioned ledger sheets - a
# different shape from the fixed-4-column helpers above (_sec_hdr/_data_row/
# _total_row), so these get their own small helper set rather than forcing
# those to fit.

def _wide_hdr_row(ws, row, org_name, title, n_cols):
    ws.merge_cells(f'A{row}:{get_column_letter(n_cols)}{row}')
    c = ws[f'A{row}']
    c.value = f'{org_name.upper()}  -  {title}'
    c.font = Font(name='Arial', bold=True, size=13, color=WHITE)
    c.fill = NAVY_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 26
    return row + 1


def _wide_col_hdrs(ws, row, labels):
    for i, lbl in enumerate(labels):
        c = ws.cell(row=row, column=i + 1, value=lbl)
        c.font = SUBHDR_FONT; c.fill = TEAL_FILL; c.border = THIN_BORDER
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 18
    return row + 1


def _month_band(ws, row, month_label, n_cols):
    ws.merge_cells(f'A{row}:{get_column_letter(n_cols)}{row}')
    c = ws[f'A{row}']
    c.value = month_label.split()[0].upper()
    c.font = Font(name='Arial', bold=True, size=10)
    c.fill = GOLD_FILL
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[row].height = 16
    return row + 1


def _payout_band(ws, row, label, total, n_cols, amount_col):
    for i in range(n_cols):
        c = ws.cell(row=row, column=i + 1)
        c.fill = LTBLUE_FILL; c.border = THIN_BORDER
    ws.cell(row=row, column=1, value=label).font = BOLD_FONT
    ws.cell(row=row, column=1).alignment = Alignment(indent=1)
    c = ws.cell(row=row, column=amount_col + 1, value=total)
    c.font = BOLD_FONT; c.number_format = MONEY_FMT
    c.alignment = Alignment(horizontal='right')
    ws.row_dimensions[row].height = 16
    return row + 1


def _deposit_band(ws, row, date_str, total, month_short, bank_stmt_short, amount_col):
    """
    Header row for a group of same-date Credits rows that together match one
    confirmed real Chase bank deposit (see build_credits_sheet) - shows the
    deposit as a single auditable total, with its QuickBooks category
    breakdown (and each line's payee) nested underneath via plain
    _wide_data_row calls. PAYEE is left blank on the band row itself since
    a single bank deposit can combine lines from different payers - see
    each nested row for who each portion came from.
    """
    for i in range(8):
        c = ws.cell(row=row, column=i + 1)
        c.fill = LTBLUE_FILL; c.border = GROUP_TOP_BORDER
    ws.cell(row=row, column=1, value=date_str).font = BOLD_FONT
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', indent=1)
    ws.cell(row=row, column=3, value='Bank Deposit').font = BOLD_FONT
    ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    c = ws.cell(row=row, column=amount_col + 1, value=total)
    c.font = BOLD_FONT; c.number_format = MONEY_FMT; c.alignment = Alignment(horizontal='right')
    ws.cell(row=row, column=5, value=month_short).font = BOLD_FONT
    ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
    ws.cell(row=row, column=6, value=bank_stmt_short).font = BOLD_FONT
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 16
    return row + 1


def _wide_data_row(ws, row, values, money_cols=(), border=None):
    border = border or THIN_BORDER
    for i, val in enumerate(values):
        c = ws.cell(row=row, column=i + 1, value=val)
        c.font = BODY_FONT; c.border = border
        if i in money_cols:
            c.number_format = MONEY_FMT
            c.alignment = Alignment(horizontal='right')
        elif i == 0:
            c.alignment = Alignment(horizontal='left', indent=1)
        else:
            c.alignment = Alignment(horizontal='center')
    ws.row_dimensions[row].height = 15
    return row + 1


def _wide_total_row(ws, row, n_cols, label, val, amount_col):
    for i in range(n_cols):
        c = ws.cell(row=row, column=i + 1)
        c.fill = LTBLUE_FILL; c.border = MED_BORDER
    ws.cell(row=row, column=1, value=label).font = TOTAL_FONT
    ws.cell(row=row, column=1).alignment = Alignment(indent=1)
    c = ws.cell(row=row, column=amount_col + 1, value=val)
    c.font = TOTAL_FONT; c.number_format = MONEY_FMT
    c.alignment = Alignment(horizontal='right')
    ws.row_dimensions[row].height = 18
    return row + 1


def _sort_by_date(txns):
    def key(t):
        try:
            return datetime.strptime(t['date'], '%m/%d/%Y')
        except (ValueError, TypeError):
            return datetime.min
    return sorted(txns, key=key)


def build_credits_sheet(ws, credits_by_month, org_name, qb_to_budget_map=None):
    """
    credits_by_month: list of (month_label, [transaction, ...]) in chronological
    fiscal-year order. Each transaction is one dict from
    parsers.parse_quickbooks_detail()['transactions'] where is_income is True.

    Only auto-derivable columns are populated (date, payee, category, amount,
    month, mapped budget line, running total) - there's no Notes column here
    because nothing in the parsed data maps to hand-written reconciliation
    notes. PAYEE comes straight from QuickBooks' own Name field (parsers.py)
    - often blank for electronic Givebacks/MemberHub payouts (no individual
    payer recorded), populated for a check or a named cash/check deposit -
    included so an auditor can see who a given line came from without
    cross-referencing QuickBooks by hand.

    A transaction QuickBooks recorded as a literal cash/check bank deposit
    (raw description 'DEPOSIT' or 'DEPOSIT ID NUMBER ...', not an electronic
    Givebacks/MemberHub payout) shows 'Bank Deposit (cash/check)' in CATEGORY
    instead of its QuickBooks category name (e.g. 'Book Fair') - BUDGET LINE
    still maps from the real category, so the budget rollup is unaffected;
    only the CATEGORY label changes, to distinguish it from an electronic
    payout at a glance.

    When several same-date rows together match one confirmed real Chase
    deposit (parsers.match_credits_to_bank_statement already verified the
    date-group's total against an actual bank deposit amount - e.g. a single
    $7,118.12 deposit that QuickBooks split into 'Book Fair' $6,219.40 and
    'Spiritwear' $898.72), they render as a bold 'Bank Deposit' band showing
    the real deposit total, with the QuickBooks category breakdown (and each
    line's own payee) nested underneath - an auditor matching against the
    Chase statement sees the deposit total first, not two unrelated-looking
    category rows that silently need to be added together to reconcile. The
    whole group (band + its nested rows) is bracketed with a bold outer
    border and thin seams between rows inside it, so it visually reads as
    one deposit, not a run of ordinary transaction rows that happen to
    follow each other. A lone row, or a group that never matched a real
    bank deposit (bank_statement_month is None - can't claim a grouping
    that isn't verified), still renders flat as before.
    """
    qb_to_budget_map = qb_to_budget_map or {}
    ws.sheet_view.showGridLines = False
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'],
                       [14, 22, 24, 14, 10, 16, 22, 16]):
        ws.column_dimensions[col].width = w

    row = _wide_hdr_row(ws, 1, org_name, 'Credits (Deposits) - Fiscal Year', 8)
    row += 1
    row = _wide_col_hdrs(row=row, ws=ws, labels=[
        'DEPOSIT DATE', 'PAYEE', 'CATEGORY', '$ AMOUNT', 'MONTH', 'BANK STATEMENT',
        'BUDGET LINE', 'RUNNING TOTAL'])

    def render_flat_row(t, running):
        budget_line = qb_to_budget_map.get(t['category'], t['category'])
        desc = (t.get('description') or '').strip().upper()
        is_bank_deposit = desc == 'DEPOSIT' or desc.startswith('DEPOSIT ID NUMBER')
        category_label = 'Bank Deposit (cash/check)' if is_bank_deposit else t['category']
        bank_stmt = t.get('bank_statement_month')
        bank_stmt = bank_stmt.split()[0] if bank_stmt else '—'
        return [t['date'], t.get('payee', ''), category_label, t['amount'],
                month_label.split()[0], bank_stmt, budget_line, running]

    running = 0.0
    for month_label, txns in credits_by_month:
        if not txns:
            continue
        row = _month_band(ws, row, month_label, n_cols=8)

        by_date = {}
        for t in _sort_by_date(txns):
            by_date.setdefault(t['date'], []).append(t)

        for date_str, group in by_date.items():
            bank_stmt_val = group[0].get('bank_statement_month')
            if len(group) > 1 and bank_stmt_val:
                deposit_total = round(sum(t['amount'] for t in group), 2)
                row = _deposit_band(ws, row, date_str, deposit_total,
                                     month_label.split()[0], bank_stmt_val.split()[0],
                                     amount_col=3)
                for i, t in enumerate(group):
                    running += t['amount']
                    budget_line = qb_to_budget_map.get(t['category'], t['category'])
                    border = GROUP_LAST_BORDER if i == len(group) - 1 else GROUP_MID_BORDER
                    row = _wide_data_row(ws, row, [
                        None, t.get('payee', ''), t['category'], t['amount'], None, None,
                        budget_line, running,
                    ], money_cols={3, 7}, border=border)
            else:
                for t in group:
                    running += t['amount']
                    row = _wide_data_row(ws, row, render_flat_row(t, running), money_cols={3, 7})

    row = _wide_total_row(ws, row, 8, 'TOTAL CREDITS', running, amount_col=3)
    return row


def build_debits_sheet(ws, debits_by_month, org_name, qb_to_budget_map=None):
    """
    debits_by_month: list of (month_label, [transaction, ...]) in chronological
    fiscal-year order. Each transaction is one dict from
    parsers.parse_quickbooks_detail()['transactions'] where is_income is False.

    NOTES column is intentionally left blank - the source data has no field
    for hand-written reconciliation notes (e.g. "check was signed late",
    explanations for corrections); fill those in by hand after generating.
    """
    qb_to_budget_map = qb_to_budget_map or {}
    ws.sheet_view.showGridLines = False
    for col, w in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I'],
                       [10, 12, 22, 26, 12, 10, 16, 22, 24]):
        ws.column_dimensions[col].width = w

    row = _wide_hdr_row(ws, 1, org_name, 'Debits (Checks/Expenses) - Fiscal Year', 9)
    row += 1
    row = _wide_col_hdrs(row=row, ws=ws, labels=[
        'CHECK #', 'DATE', 'PAYEE', 'CATEGORY', '$ AMOUNT',
        'MONTH', 'BANK STATEMENT', 'BUDGET LINE', 'NOTES'])
    # Running total lives in a 10th column, added after NOTES so a blank
    # Notes cell doesn't visually break the money column next to it.
    ws.cell(row=row - 1, column=10, value='RUNNING TOTAL').font = SUBHDR_FONT
    ws.cell(row=row - 1, column=10).fill = TEAL_FILL
    ws.cell(row=row - 1, column=10).border = THIN_BORDER
    ws.cell(row=row - 1, column=10).alignment = Alignment(horizontal='center', vertical='center')
    ws.column_dimensions['J'].width = 16

    running = 0.0
    for month_label, txns in debits_by_month:
        if not txns:
            continue
        row = _month_band(ws, row, month_label, n_cols=10)
        for t in _sort_by_date(txns):
            running += t['amount']
            budget_line = qb_to_budget_map.get(t['category'], t['category'])
            bank_stmt = t.get('bank_statement_month')
            bank_stmt = bank_stmt.split()[0] if bank_stmt else '—'
            row = _wide_data_row(ws, row, [
                t['check_no'], t['date'], t['payee'], t['category'], t['amount'],
                month_label.split()[0], bank_stmt, budget_line, '', running,
            ], money_cols={4, 9})

    row = _wide_total_row(ws, row, 10, 'TOTAL DEBITS', running, amount_col=4)
    return row


def build_memberhub_summary_sheet(ws, givebacks_by_month, org_name):
    """
    givebacks_by_month: list of (month_label, [payout, ...]) in chronological
    fiscal-year order. Each payout is a dict {date, total, items} - date is
    the matched bank deposit date (None if unreconciled - rendered as such,
    never dropped), items is that one payout file's line items (dicts with
    item/category/count/total/source_file, from
    parsers.parse_givebacks_files() called once per payout file). Payouts
    within a month are sorted by date, unreconciled ones listed last.
    """
    ws.sheet_view.showGridLines = False
    for col, w in zip(['A', 'B', 'C', 'D'], [30, 24, 14, 16]):
        ws.column_dimensions[col].width = w

    row = _wide_hdr_row(ws, 1, org_name, 'MemberHub / Givebacks Summary - Fiscal Year', 4)
    row += 1
    row = _wide_col_hdrs(row=row, ws=ws, labels=[
        'ITEM', 'CATEGORY', '$ AMOUNT', 'RUNNING TOTAL'])

    running = 0.0
    for month_label, payouts in givebacks_by_month:
        if not payouts:
            continue
        row = _month_band(ws, row, month_label, n_cols=4)
        sorted_payouts = sorted(payouts, key=lambda p: (p['date'] is None, p['date'] or ''))
        for payout in sorted_payouts:
            label = f"Payout — {payout['date']}" if payout['date'] else 'Payout — (date not reconciled)'
            row = _payout_band(ws, row, label, payout['total'], n_cols=4, amount_col=2)
            for it in payout['items']:
                running += it['total']
                row = _wide_data_row(ws, row, [
                    it['item'], it['category'], it['total'], running,
                ], money_cols={2, 3})

    row = _wide_total_row(ws, row, 4, 'TOTAL GIVEBACKS', running, amount_col=2)
    return row
